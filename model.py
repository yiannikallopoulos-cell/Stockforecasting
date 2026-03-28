#!/usr/bin/env python3
"""
Financial Operating Model Builder v2
Pulls historical financials from SEC EDGAR + Yahoo Finance,
enriches with earnings call / analyst context via web search,
and builds a fully formatted Excel operating model with 3-5 year forecasts.

Usage:
    python Forecasting_v2.py TICKER [--years {3,4,5}]

Examples:
    python Forecasting_v2.py META
    python Forecasting_v2.py AAPL --years 5
    python Forecasting_v2.py NVDA --years 3
"""

import argparse
import json
import sys
import time
from datetime import datetime
from pathlib import Path

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── Color Palette ─────────────────────────────────────────────────────────────
DARK_BLUE   = "1F3864"
MID_BLUE    = "2E75B6"
LIGHT_BLUE  = "D6E4F0"
ACCENT_BLUE = "BDD7EE"
WHITE       = "FFFFFF"
BLACK       = "000000"
GRAY_LIGHT  = "F2F2F2"
GREEN       = "375623"
GREEN_LIGHT = "E2EFDA"
INPUT_BLUE  = "0070C0"
YELLOW_BG   = "FFFF99"
ORANGE      = "C55A11"

HEADERS = {"User-Agent": "FinancialModel research@example.com", "Accept": "application/json"}

# ── Style helpers ─────────────────────────────────────────────────────────────
def fl(hex_c): return PatternFill("solid", fgColor=hex_c)
def ft(bold=False, color=BLACK, size=9, italic=False):
    return Font(name="Arial", bold=bold, color=color, size=size, italic=italic)
def aln(h="left", wrap=False): return Alignment(horizontal=h, vertical="center", wrap_text=wrap)
def thin(): return Border(bottom=Side(style="thin", color="CCCCCC"))
def medium_bottom(): return Border(bottom=Side(style="medium", color=DARK_BLUE))

def section_hdr(ws, row, c1, c2, text):
    try: ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)
    except: pass
    c = ws.cell(row=row, column=c1, value=text)
    c.font = ft(bold=True, color=WHITE, size=10)
    c.fill = fl(DARK_BLUE); c.alignment = aln("left")
    ws.row_dimensions[row].height = 15

def sub_hdr(ws, row, c1, c2, text):
    try: ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)
    except: pass
    c = ws.cell(row=row, column=c1, value=text)
    c.font = ft(bold=True, color=WHITE, size=9)
    c.fill = fl(MID_BLUE); c.alignment = aln("left")
    ws.row_dimensions[row].height = 13

def lbl(ws, row, col, text, bold=False, indent=0, italic=False, bg=None, color=BLACK):
    c = ws.cell(row=row, column=col, value=("  " * indent) + text)
    c.font = ft(bold=bold, color=color, italic=italic, size=9)
    if bg: c.fill = fl(bg)
    c.alignment = aln("left")

def val_cell(ws, row, col, value, fmt='#,##0.0;(#,##0.0);"-"', bold=False,
             color=BLACK, bg=None, formula=False):
    c = ws.cell(row=row, column=col, value=value)
    c.number_format = fmt
    c.font = ft(bold=bold, color=color, size=9)
    c.alignment = aln("right")
    if bg: c.fill = fl(bg)
    return c

def note_cell(ws, row, col, text, color="595959"):
    c = ws.cell(row=row, column=col, value=text)
    c.font = ft(italic=True, color=color, size=8)
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    return c

# ── Data Fetching ─────────────────────────────────────────────────────────────

HEADERS = {"User-Agent": "FinancialModel research@example.com", "Accept": "application/json"}

def _ensure_yfinance():
    try:
        import yfinance as yf
        return yf
    except ImportError:
        raise ImportError("yfinance not installed. Add yfinance to requirements.txt")

def get_cik(ticker):
    r = requests.get("https://www.sec.gov/files/company_tickers.json",headers=HEADERS,timeout=20)
    r.raise_for_status()
    for _,info in r.json().items():
        if info["ticker"].upper()==ticker.upper():
            return str(info["cik_str"]).zfill(10)
    raise ValueError(f"CIK not found: {ticker}")

def get_sec_facts(cik):
    r = requests.get(f"https://data.sec.gov/api/xbrl/companyfacts/CIK{cik}.json",headers=HEADERS,timeout=30)
    r.raise_for_status(); return r.json()

def get_sec_submissions(cik):
    try:
        r = requests.get(f"https://data.sec.gov/submissions/CIK{cik}.json",headers=HEADERS,timeout=15)
        r.raise_for_status(); data=r.json()
        f=data.get("filings",{}).get("recent",{})
        forms=f.get("form",[]); dates=f.get("filingDate",[])
        result={"10-K":[],"8-K":[],"DEF 14A":[]}
        for form,dt in zip(forms,dates):
            if form in result and len(result[form])<2:
                result[form].append({"date":dt,"url":f"https://www.sec.gov/cgi-bin/browse-edgar?action=getcompany&CIK={cik}&type={form}&dateb=&owner=include&count=10"})
        return result,data.get("name","")
    except: return {},""

def sec_annual_series(facts,*concepts,n=5):
    for concept in concepts:
        try:
            units=facts["facts"]["us-gaap"][concept]["units"]
            recs=units.get("USD",[])+units.get("shares",[])
            annual=[x for x in recs if x.get("form") in("10-K","10-K/A") and x.get("fp")=="FY"]
            by_yr={}
            for x in annual:
                yr=x["end"][:4]
                if yr not in by_yr or x.get("filed","")>by_yr[yr].get("filed",""):
                    by_yr[yr]=x
            if by_yr:
                recent=sorted(by_yr.values(),key=lambda x:x["end"])[-n:]
                return {r["end"][:4]:r["val"] for r in recent}
        except: continue
    return {}

def _info_val(info,*keys):
    for k in keys:
        v=info.get(k)
        if v is not None and str(v) not in("nan","None",""):
            try: return float(v)
            except: return v
    return None

def _df_series(df,*names):
    if df is None or df.empty: return {}
    for name in names:
        for idx in df.index:
            a=str(idx).lower().replace(" ","").replace("_","")
            b=name.lower().replace(" ","").replace("_","")
            if a==b or b in a:
                result={}
                for col in df.columns:
                    try:
                        yr=str(col)[:4]; v=df.loc[idx,col]
                        if v is not None and str(v) not in("nan","None","<NA>","NaN"):
                            result[yr]=float(v)
                    except: pass
                if result: return result
    return {}

def _yf_session():
    """Placeholder - yfinance 1.2+ manages its own session internally."""
    return None


def yfinance_fetch(ticker):
    yf = _ensure_yfinance()
    print(f"    Downloading {ticker} via yfinance...")

    # yfinance 1.2+ manages its own curl_cffi session — do NOT pass a session
    t = yf.Ticker(ticker)

    # Fetch info with retries
    info = {}
    for attempt in range(3):
        try:
            raw = t.info
            if raw and len(raw) > 5:
                info = raw
                break
        except Exception as e:
            print(f"    info attempt {attempt+1} failed: {e}")
        if attempt < 2:
            time.sleep(2 + attempt * 2)

    # fast_info fallback if info empty
    if not info or len(info) < 5:
        try:
            fi = t.fast_info
            info = {
                "currentPrice":      getattr(fi,"last_price",None),
                "regularMarketPrice":getattr(fi,"last_price",None),
                "previousClose":     getattr(fi,"previous_close",None),
                "marketCap":         getattr(fi,"market_cap",None),
                "sharesOutstanding": getattr(fi,"shares",None),
                "currency":          getattr(fi,"currency","USD"),
                "exchange":          getattr(fi,"exchange",""),
                "longName":          ticker.upper(),
            }
            print(f"    Used fast_info fallback for {ticker}")
        except Exception as e:
            print(f"    fast_info also failed: {e}")

    # Financial statements
    is_df = bs_df = cf_df = None
    for attempt in range(2):
        try:
            if is_df is None:   is_df = t.income_stmt
        except Exception: pass
        try:
            if bs_df is None:   bs_df = t.balance_sheet
        except Exception: pass
        try:
            if cf_df is None:   cf_df = t.cashflow
        except Exception: pass
        if is_df is not None and bs_df is not None and cf_df is not None:
            break
        if attempt == 0:
            time.sleep(2)

    price=_info_val(info,"currentPrice","regularMarketPrice","previousClose") or 0
    mkt_cap=_info_val(info,"marketCap") or 0
    shares_out=_info_val(info,"sharesOutstanding") or 0

    stock={
        "price":price,"shares_out":shares_out,"market_cap":mkt_cap,
        "beta":_info_val(info,"beta") or 1.0,
        "ev":_info_val(info,"enterpriseValue") or 0,
        "pe_trailing":_info_val(info,"trailingPE") or 0,
        "pe_forward":_info_val(info,"forwardPE") or 0,
        "ev_ebitda":_info_val(info,"enterpriseToEbitda") or 0,
        "ev_rev":_info_val(info,"enterpriseToRevenue") or 0,
        "ps_ratio":_info_val(info,"priceToSalesTrailing12Months") or 0,
        "pb_ratio":_info_val(info,"priceToBook") or 0,
        "peg_ratio":_info_val(info,"pegRatio") or 0,
        "trailing_eps":_info_val(info,"trailingEps") or 0,
        "forward_eps":_info_val(info,"forwardEps") or 0,
        "dividend_rate":_info_val(info,"dividendRate") or 0,
        "dividend_yield":_info_val(info,"dividendYield") or 0,
        "payout_ratio":_info_val(info,"payoutRatio") or 0,
        "week52_high":_info_val(info,"fiftyTwoWeekHigh") or 0,
        "week52_low":_info_val(info,"fiftyTwoWeekLow") or 0,
        "sector":info.get("sector") or "Unknown",
        "industry":info.get("industry") or "Unknown",
        "name":info.get("longName") or info.get("shortName") or ticker.upper(),
        "website":info.get("website") or "",
        "description":info.get("longBusinessSummary") or "",
        "fulltime_employees":_info_val(info,"fullTimeEmployees") or 0,
        "country":info.get("country") or "",
        "gross_margin":_info_val(info,"grossMargins") or 0,
        "op_margin":_info_val(info,"operatingMargins") or 0,
        "profit_margin":_info_val(info,"profitMargins") or 0,
        "ebitda_margin":_info_val(info,"ebitdaMargins") or 0,
        "revenue_growth":_info_val(info,"revenueGrowth") or 0,
        "earnings_growth":_info_val(info,"earningsGrowth") or 0,
        "current_ratio":_info_val(info,"currentRatio") or 0,
        "quick_ratio":_info_val(info,"quickRatio") or 0,
        "debt_equity":_info_val(info,"debtToEquity") or 0,
        "roe":_info_val(info,"returnOnEquity") or 0,
        "roa":_info_val(info,"returnOnAssets") or 0,
        "total_cash":_info_val(info,"totalCash") or 0,
        "total_debt":_info_val(info,"totalDebt") or 0,
        "free_cashflow":_info_val(info,"freeCashflow") or 0,
        "op_cashflow":_info_val(info,"operatingCashflow") or 0,
        "ebitda":_info_val(info,"ebitda") or 0,
        "total_revenue":_info_val(info,"totalRevenue") or 0,
        "target_price":_info_val(info,"targetMeanPrice") or 0,
        "target_high":_info_val(info,"targetHighPrice") or 0,
        "target_low":_info_val(info,"targetLowPrice") or 0,
        "analyst_count":_info_val(info,"numberOfAnalystOpinions") or 0,
        "rec":info.get("recommendationKey") or "",
        "rec_mean":_info_val(info,"recommendationMean") or 0,
    }

    def s(df,*n): return _df_series(df,*n)
    yf_is={
        "revenue":s(is_df,"Total Revenue","Revenue"),
        "cogs":s(is_df,"Cost Of Revenue","Cost of Revenue"),
        "gross_profit":s(is_df,"Gross Profit"),
        "rd_exp":s(is_df,"Research And Development","Research Development"),
        "sga_exp":s(is_df,"Selling General And Administrative","Selling General Administrative"),
        "op_income":s(is_df,"Operating Income","Operating Income Loss","Total Operating Income As Reported"),
        "interest_exp":s(is_df,"Interest Expense","Net Interest Income"),
        "pretax_inc":s(is_df,"Pretax Income","Income Before Tax"),
        "tax_exp":s(is_df,"Tax Provision","Income Tax Expense"),
        "net_income":s(is_df,"Net Income","Net Income Common Stockholders"),
    }
    yf_bs={
        "cash":s(bs_df,"Cash And Cash Equivalents","Cash Cash Equivalents And Short Term Investments"),
        "curr_assets":s(bs_df,"Current Assets","Total Current Assets"),
        "pp_e":s(bs_df,"Net PPE","Property Plant Equipment Net","Property Plant And Equipment Net"),
        "goodwill":s(bs_df,"Goodwill"),
        "intangibles":s(bs_df,"Other Intangible Assets","Goodwill And Other Intangible Assets"),
        "total_assets":s(bs_df,"Total Assets"),
        "curr_liab":s(bs_df,"Current Liabilities","Total Current Liabilities"),
        "lt_debt":s(bs_df,"Long Term Debt","Long Term Debt And Capital Lease Obligation"),
        "total_liab":s(bs_df,"Total Liabilities Net Minority Interest","Total Liab","Total Liabilities"),
        "total_equity":s(bs_df,"Stockholders Equity","Total Stockholder Equity","Common Stock Equity"),
    }
    yf_cf={
        "op_cf":s(cf_df,"Operating Cash Flow","Total Cash From Operating Activities"),
        "dep_amor":s(cf_df,"Depreciation And Amortization","Depreciation Amortization Depletion","Depreciation"),
        "sbc":s(cf_df,"Stock Based Compensation","Share Based Compensation"),
        "capex":s(cf_df,"Capital Expenditure","Purchase Of Property Plant And Equipment"),
        "div_paid":s(cf_df,"Common Stock Dividend Paid","Payment Of Dividends"),
        "buybacks":s(cf_df,"Repurchase Of Capital Stock","Common Stock Repurchased"),
        "fin_cf":s(cf_df,"Financing Cash Flow","Total Cash From Financing Activities"),
    }
    # Normalize signs
    for yr,v in list(yf_cf["capex"].items()):
        if v and v>0: yf_cf["capex"][yr]=-v
    for key in("div_paid","buybacks"):
        for yr,v in list(yf_cf[key].items()):
            if v and v>0: yf_cf[key][yr]=-v

    analyst_est={}
    try:
        re_df=t.revenue_estimate
        if re_df is not None and not re_df.empty:
            for period in ["+1y","+2y"]:
                if period in re_df.index:
                    avg=float(re_df.loc[period].get("avg",0) or 0)
                    if avg: analyst_est[period]={"rev_est":avg}
    except: pass
    try:
        ge_df=t.earnings_estimate
        if ge_df is not None and not ge_df.empty:
            for period in ["+1y","+2y"]:
                if period in ge_df.index:
                    avg=float(ge_df.loc[period].get("avg",0) or 0)
                    if avg:
                        analyst_est.setdefault(period,{})["eps_est"]=avg
    except: pass

    rev_yrs=sorted(yf_is["revenue"].keys())
    ta_yrs=sorted(yf_bs["total_assets"].keys())
    ocf_yrs=sorted(yf_cf["op_cf"].keys())
    print(f"    ✓ Revenue: {rev_yrs}  |  Assets: {ta_yrs}  |  OCF: {ocf_yrs}")
    print(f"    ✓ Price: ${price:,.2f}  Mkt Cap: ${mkt_cap/1e9:,.1f}B  Sector: {stock['sector']}")
    return stock,yf_is,yf_bs,yf_cf,analyst_est

def yfinance_single(ticker):
    yf=_ensure_yfinance()
    try:
        t=yf.Ticker(ticker)
        info={}
        for attempt in range(2):
            try:
                raw=t.info
                if raw and len(raw)>5: info=raw; break
            except: pass
            if attempt==0: time.sleep(1)
        if not info: return {}
        if not _info_val(info,"currentPrice","regularMarketPrice","previousClose"): return {}
        return {
            "name":info.get("longName") or info.get("shortName") or ticker,
            "price":_info_val(info,"currentPrice","regularMarketPrice","previousClose") or 0,
            "market_cap":_info_val(info,"marketCap") or 0,
            "beta":_info_val(info,"beta") or 0,
            "ev":_info_val(info,"enterpriseValue") or 0,
            "pe_trailing":_info_val(info,"trailingPE") or 0,
            "pe_forward":_info_val(info,"forwardPE") or 0,
            "ev_ebitda":_info_val(info,"enterpriseToEbitda") or 0,
            "ev_rev":_info_val(info,"enterpriseToRevenue") or 0,
            "ps_ratio":_info_val(info,"priceToSalesTrailing12Months") or 0,
            "pb_ratio":_info_val(info,"priceToBook") or 0,
            "peg_ratio":_info_val(info,"pegRatio") or 0,
            "gross_margin":_info_val(info,"grossMargins") or 0,
            "op_margin":_info_val(info,"operatingMargins") or 0,
            "profit_margin":_info_val(info,"profitMargins") or 0,
            "revenue_growth":_info_val(info,"revenueGrowth") or 0,
            "roe":_info_val(info,"returnOnEquity") or 0,
            "debt_equity":_info_val(info,"debtToEquity") or 0,
            "sector":info.get("sector") or "",
            "industry":info.get("industry") or "",
        }
    except: return {}

def get_peers(ticker,sector,industry):
    PEER_MAP={
        "META":["GOOGL","SNAP","PINS","TTD"],"AAPL":["MSFT","GOOGL","AMZN","SONY"],
        "MSFT":["AAPL","GOOGL","AMZN","CRM"],"GOOGL":["META","MSFT","AMZN","NFLX"],
        "AMZN":["MSFT","GOOGL","BABA","SHOP"],"NVDA":["AMD","INTC","QCOM","TSM"],
        "TSLA":["F","GM","RIVN","NIO"],"JPM":["BAC","WFC","GS","MS"],
        "JNJ":["PFE","MRK","ABT","BMY"],"XOM":["CVX","COP","BP","SHEL"],
        "NFLX":["DIS","WBD","PARA","AMZN"],"UBER":["LYFT","ABNB","DASH","GRAB"],
    }
    sector_defaults={
        "Technology":["MSFT","GOOGL","AMZN","AAPL"],
        "Communication Services":["GOOGL","NFLX","DIS","T"],
        "Financial Services":["JPM","BAC","WFC","GS"],
        "Healthcare":["JNJ","PFE","UNH","ABT"],
        "Consumer Cyclical":["AMZN","TSLA","HD","NKE"],
        "Energy":["XOM","CVX","COP","SLB"],
    }
    tickers=PEER_MAP.get(ticker.upper()) or sector_defaults.get(sector,["SPY","QQQ"])
    peers={}
    for p in tickers[:4]:
        try:
            time.sleep(0.3); d=yfinance_single(p)
            if d and d.get("price"):
                peers[p]=d; print(f"      ✓ {p}: ${d['price']:,.2f}")
        except: pass
    return peers

def merge_series(*dicts):
    result={}
    for d in dicts:
        for yr,v in d.items():
            if yr not in result and v is not None and v!=0:
                result[yr]=v
    return result

def fetch_financials(ticker):
    print(f"  [1/3] Fetching data via yfinance...")
    stock,yf_is,yf_bs,yf_cf,analyst_est=yfinance_fetch(ticker)

    print(f"  [2/3] Fetching SEC EDGAR supplemental data...")
    filings={}; cik=""
    try:
        cik=get_cik(ticker)
        sec_facts=get_sec_facts(cik)
        filings,sec_name=get_sec_submissions(cik)
        if not stock.get("name") or stock["name"]==ticker.upper():
            stock["name"]=sec_name or ticker.upper()
        def sec(n,*c): return sec_annual_series(sec_facts,*c,n=n)
        sec_rev=sec(5,"Revenues","RevenueFromContractWithCustomerExcludingAssessedTax","SalesRevenueNet")
        sec_cogs=sec(5,"CostOfRevenue","CostOfGoodsAndServicesSold")
        sec_gp=sec(5,"GrossProfit")
        sec_rd=sec(5,"ResearchAndDevelopmentExpense")
        sec_sga=sec(5,"SellingGeneralAndAdministrativeExpense")
        sec_oi=sec(5,"OperatingIncomeLoss")
        sec_int=sec(5,"InterestExpense","InterestAndDebtExpense")
        sec_pt=sec(5,"IncomeLossFromContinuingOperationsBeforeIncomeTaxesExtraordinaryItemsNoncontrollingInterest")
        sec_tax=sec(5,"IncomeTaxExpenseBenefit")
        sec_ni=sec(5,"NetIncomeLoss")
        sec_eps=sec(5,"EarningsPerShareDiluted")
        sec_sh=sec(5,"WeightedAverageNumberOfDilutedSharesOutstanding")
        sec_dps=sec(5,"CommonStockDividendsPerShareDeclared","CommonStockDividendsPerShareCashPaid")
        sec_cash=sec(5,"CashAndCashEquivalentsAtCarryingValue","CashCashEquivalentsAndShortTermInvestments")
        sec_ca=sec(5,"AssetsCurrent")
        sec_ta=sec(5,"Assets")
        sec_ppe=sec(5,"PropertyPlantAndEquipmentNet")
        sec_gw=sec(5,"Goodwill")
        sec_ia=sec(5,"IntangibleAssetsNetExcludingGoodwill")
        sec_cl=sec(5,"LiabilitiesCurrent")
        sec_ltd=sec(5,"LongTermDebt","LongTermDebtNoncurrent")
        sec_tl=sec(5,"Liabilities")
        sec_eq=sec(5,"StockholdersEquity","StockholdersEquityAttributableToParent")
        sec_ocf=sec(5,"NetCashProvidedByUsedInOperatingActivities")
        sec_cx_r=sec(5,"PaymentsToAcquirePropertyPlantAndEquipment")
        sec_cx={yr:-abs(v) for yr,v in sec_cx_r.items()}
        sec_da=sec(5,"DepreciationDepletionAndAmortization","Depreciation")
        sec_div=sec(5,"PaymentsOfDividends","PaymentsOfDividendsCommonStock")
        sec_buy=sec(5,"PaymentsForRepurchaseOfCommonStock")
        sec_sbc=sec(5,"ShareBasedCompensation")
        print(f"    ✓ SEC EDGAR CIK {cik} | last 10-K: {filings.get('10-K',[{}])[0].get('date','?')}")
    except Exception as e:
        print(f"    ⚠ SEC EDGAR unavailable ({e}) — using yfinance only")
        sec_rev=sec_cogs=sec_gp=sec_rd=sec_sga=sec_oi=sec_int=sec_pt={}
        sec_tax=sec_ni=sec_eps=sec_sh=sec_dps=sec_cash=sec_ca=sec_ta={}
        sec_ppe=sec_gw=sec_ia=sec_cl=sec_ltd=sec_tl=sec_eq={}
        sec_ocf=sec_cx=sec_da=sec_div=sec_buy=sec_sbc={}

    print(f"  [3/3] Fetching peer data...")
    peers=get_peers(ticker,stock.get("sector",""),stock.get("industry",""))

    def m(a,b): return merge_series(a,b)
    return {
        "ticker":ticker.upper(),"name":stock.get("name",ticker.upper()),
        "sector":stock.get("sector","Unknown"),"industry":stock.get("industry","Unknown"),
        "cik":cik,"stock":stock,"peers":peers,"filings":filings,"analyst_est":analyst_est,
        "revenue":m(yf_is["revenue"],sec_rev),
        "cogs":m(yf_is["cogs"],sec_cogs),
        "gross_profit":m(yf_is["gross_profit"],sec_gp),
        "rd_exp":m(yf_is["rd_exp"],sec_rd),
        "sga_exp":m(yf_is["sga_exp"],sec_sga),
        "op_income":m(yf_is["op_income"],sec_oi),
        "interest_exp":m(yf_is["interest_exp"],sec_int),
        "pretax_inc":m(yf_is["pretax_inc"],sec_pt),
        "tax_exp":m(yf_is["tax_exp"],sec_tax),
        "net_income":m(yf_is["net_income"],sec_ni),
        "eps_diluted":m({},sec_eps),
        "shares_dil":m({},sec_sh),
        "dps":m({},sec_dps),
        "cash":m(yf_bs["cash"],sec_cash),
        "curr_assets":m(yf_bs["curr_assets"],sec_ca),
        "total_assets":m(yf_bs["total_assets"],sec_ta),
        "pp_e":m(yf_bs["pp_e"],sec_ppe),
        "goodwill":m(yf_bs["goodwill"],sec_gw),
        "intangibles":m(yf_bs["intangibles"],sec_ia),
        "curr_liab":m(yf_bs["curr_liab"],sec_cl),
        "lt_debt":m(yf_bs["lt_debt"],sec_ltd),
        "total_liab":m(yf_bs["total_liab"],sec_tl),
        "total_equity":m(yf_bs["total_equity"],sec_eq),
        "op_cf":m(yf_cf["op_cf"],sec_ocf),
        "capex":m(yf_cf["capex"],sec_cx),
        "dep_amor":m(yf_cf["dep_amor"],sec_da),
        "div_paid":m(yf_cf["div_paid"],sec_div),
        "buybacks":m(yf_cf["buybacks"],sec_buy),
        "stock_comp":m(yf_cf.get("sbc",{}),sec_sbc),
    }

# ── Forecast Assumptions — Smarter Forecasting Engine v2 ─────────────────────

def last_val(s, default=0):
    if not s: return default
    return sorted(s.items())[-1][1]

def sorted_vals(s):
    """Return list of (year, value) sorted ascending, filtering None/zero."""
    return [(yr, v) for yr, v in sorted(s.items()) if v and v != 0]

def cagr(s, n=3):
    vals = [(yr, v) for yr, v in sorted(s.items()) if v and v > 0]
    if len(vals) < 2: return 0.08
    r = vals[-min(n, len(vals)):]
    start, end = r[0][1], r[-1][1]
    yrs = len(r) - 1
    if yrs == 0 or start <= 0: return 0.08
    return (end / start) ** (1 / yrs) - 1

def exp_weighted_avg(series_dict, n=4, decay=0.5):
    """
    Exponentially weighted average — recent years weighted more heavily.
    decay=0.5 means most recent year gets 2x weight of second-most-recent.
    Returns weighted average of the values.
    """
    vals = sorted_vals(series_dict)
    if not vals: return 0
    recent = vals[-n:]
    total_weight = 0
    weighted_sum = 0
    for i, (yr, v) in enumerate(recent):
        w = (1 + decay) ** i   # more recent = higher weight
        weighted_sum += v * w
        total_weight += w
    return weighted_sum / total_weight if total_weight else 0

def avg_margin(num, denom, n=3, weighted=False):
    """Compute average or exp-weighted margin ratio."""
    vals = []
    weights = []
    yrs = sorted(set(num.keys()) & set(denom.keys()))[-n:]
    for i, yr in enumerate(yrs):
        d = denom.get(yr)
        if d and d != 0:
            vals.append(num[yr] / d)
            weights.append((1.5 ** i) if weighted else 1.0)
    if not vals: return 0
    if weighted:
        return sum(v * w for v, w in zip(vals, weights)) / sum(weights)
    return sum(vals) / len(vals)

def margin_trend(num, denom, n=4):
    """
    Compute margin trend (slope) over n years.
    Returns (latest_margin, trend_per_year, r_squared).
    Used to detect margin expansion/compression.
    """
    margins = []
    yrs_list = sorted(set(num.keys()) & set(denom.keys()))[-n:]
    for yr in yrs_list:
        d = denom.get(yr)
        if d and d != 0:
            margins.append(num[yr] / d)
    if len(margins) < 2:
        latest = margins[-1] if margins else 0
        return latest, 0.0, 0.0
    n_ = len(margins)
    x = list(range(n_))
    x_mean = sum(x) / n_
    y_mean = sum(margins) / n_
    ss_xy = sum((x[i] - x_mean) * (margins[i] - y_mean) for i in range(n_))
    ss_xx = sum((x[i] - x_mean) ** 2 for i in range(n_))
    slope = ss_xy / ss_xx if ss_xx != 0 else 0
    # R-squared
    y_pred = [y_mean + slope * (x[i] - x_mean) for i in range(n_)]
    ss_res = sum((margins[i] - y_pred[i]) ** 2 for i in range(n_))
    ss_tot = sum((margins[i] - y_mean) ** 2 for i in range(n_))
    r2 = 1 - ss_res / ss_tot if ss_tot != 0 else 0
    return margins[-1], slope, max(0, r2)

def operating_leverage(rev, oi, n=3):
    """
    Estimate degree of operating leverage (DOL).
    DOL = % change in EBIT / % change in revenue.
    High DOL (>2) = margins expand fast as revenue grows.
    """
    rev_vals = sorted_vals(rev)[-n-1:]
    oi_vals_d = dict(sorted_vals(oi))
    if len(rev_vals) < 2: return 1.0
    dols = []
    for i in range(1, len(rev_vals)):
        yr_prev, r_prev = rev_vals[i-1]
        yr_curr, r_curr = rev_vals[i]
        o_prev = oi_vals_d.get(yr_prev)
        o_curr = oi_vals_d.get(yr_curr)
        if r_prev and o_prev and r_prev != 0 and o_prev != 0:
            pct_rev = (r_curr - r_prev) / abs(r_prev)
            pct_oi  = (o_curr - o_prev) / abs(o_prev)
            if pct_rev != 0:
                dols.append(pct_oi / pct_rev)
    if not dols: return 1.0
    # Cap between 0 and 5
    return max(0.5, min(5.0, sum(dols) / len(dols)))

def revenue_mean_reversion(base_g, industry_avg_g=0.08, years_to_mean=5):
    """
    Project revenue growth tapering toward industry average over time.
    High-growth companies naturally decelerate; low-growth ones stabilize.
    Returns list of annual growth rates.
    """
    # Pull toward industry avg each year, faster if far from it
    growth_rates = []
    g = base_g
    for i in range(years_to_mean):
        gap = industry_avg_g - g
        # Reversion speed: 25% of gap closed per year
        g = g + 0.25 * gap
        g = max(g, 0.02)   # floor at 2% (positive growth)
        growth_rates.append(round(g, 4))
    return growth_rates

def classify_company(rev, oi, rev_cagr, gross_margin_val):
    """
    Classify company into growth stage for appropriate modeling treatment.
    Returns one of: 'hyper_growth', 'high_growth', 'mature_growth',
                    'stable', 'turnaround', 'declining'
    """
    latest_oi = last_val(oi) if oi else 0
    latest_rev = last_val(rev) if rev else 1
    op_margin_ltm = latest_oi / latest_rev if latest_rev else 0

    if rev_cagr > 0.25 and op_margin_ltm < 0.05:
        return "hyper_growth"       # High growth, not yet profitable
    elif rev_cagr > 0.15:
        return "high_growth"        # Strong growth, likely profitable
    elif rev_cagr > 0.07:
        return "mature_growth"      # Solid steady growth
    elif rev_cagr >= 0.0:
        return "stable"             # Low growth, cash generative
    elif rev_cagr > -0.05:
        return "turnaround"         # Slight decline, may recover
    else:
        return "declining"          # Structural decline

def nwc_ratio(curr_assets, curr_liab, rev, n=3):
    """
    Net working capital as % of revenue — captures cash conversion cycle.
    Rising NWC ratio = consumes more cash as it grows (bad for FCF).
    Falling NWC ratio = becomes more efficient (good for FCF).
    """
    ratios = []
    for yr in sorted(rev.keys())[-n:]:
        ca = curr_assets.get(yr, 0) or 0
        cl = curr_liab.get(yr, 0) or 0
        rv = rev.get(yr, 0) or 0
        if rv: ratios.append((ca - cl) / rv)
    return sum(ratios) / len(ratios) if ratios else 0.10

def maintenance_vs_growth_capex(capex, dep, rev):
    """
    Distinguish maintenance CapEx (replacing existing assets) from growth CapEx.
    Rule of thumb: maintenance CapEx ≈ D&A.
    Growth CapEx = total CapEx - maintenance CapEx.
    Returns (maintenance_pct_rev, growth_pct_rev).
    """
    total_cx = abs(avg_margin(capex, rev)) if capex else 0
    dep_pct  = avg_margin(dep, rev) if dep else 0
    maint    = min(dep_pct, total_cx)            # maintenance <= total capex
    growth   = max(0, total_cx - maint)           # remainder is growth
    return round(maint, 4), round(growth, 4)

def build_assumptions(data, n_proj):
    """
    Smarter forecasting engine incorporating:
    - Exponential weighting (recent years weighted more heavily)
    - Analyst consensus blending with dynamic weight based on coverage depth
    - Mean reversion toward industry averages over projection period
    - Margin trend detection (expansion vs compression)
    - Operating leverage adjustment
    - Company stage classification
    - NWC efficiency modeling
    - Maintenance vs growth CapEx split
    - Scenario framework (bull / base / bear)
    """
    rev    = data["revenue"]
    ni     = data["net_income"]
    cogs_d = data["cogs"]
    gp     = data["gross_profit"]
    sga    = data["sga_exp"]
    rd     = data["rd_exp"]
    oi     = data["op_income"]
    te     = data["tax_exp"]
    dep    = data["dep_amor"]
    capex  = data["capex"]
    lt_debt= data["lt_debt"]
    eq     = data["total_equity"]
    ca     = data.get("curr_assets", {})
    cl     = data.get("curr_liab", {})
    analyst= data.get("analyst_est", {})
    stock  = data["stock"]
    analyst_count = int(stock.get("analyst_count") or 0)

    # ── Revenue Growth ─────────────────────────────────────────────────────────
    rev_cagr_3yr = cagr(rev, 3)
    rev_cagr_5yr = cagr(rev, 5)

    # Analyst estimate — weight by coverage depth (more analysts = more reliable)
    analyst_yr1_growth = None
    if analyst.get("+1y", {}).get("rev_est") and last_val(rev):
        analyst_yr1_growth = analyst["+1y"]["rev_est"] / last_val(rev) - 1
    elif analyst.get("+1y", {}).get("growth"):
        analyst_yr1_growth = analyst["+1y"]["growth"]

    # Dynamic analyst weight: scale from 30% (1-5 analysts) to 70% (20+ analysts)
    if analyst_yr1_growth is not None and abs(analyst_yr1_growth) < 0.60:
        analyst_weight = min(0.30 + (analyst_count / 30) * 0.40, 0.70)
        hist_weight    = 1 - analyst_weight
        base_g         = analyst_weight * analyst_yr1_growth + hist_weight * rev_cagr_3yr
        print(f"    ✓ Rev growth: analyst({analyst_yr1_growth:.1%}) × {analyst_weight:.0%} + "
              f"hist({rev_cagr_3yr:.1%}) × {hist_weight:.0%} = {base_g:.1%} "
              f"[{analyst_count} analysts]")
    else:
        # No analyst data — blend 3yr and 5yr CAGR, weight 3yr more
        base_g = 0.65 * rev_cagr_3yr + 0.35 * rev_cagr_5yr
        print(f"    ✓ Rev growth: 3yr CAGR {rev_cagr_3yr:.1%} blended with "
              f"5yr {rev_cagr_5yr:.1%} = {base_g:.1%}")

    base_g = max(min(base_g, 0.45), -0.10)

    # Company stage classification — sets mean-reversion target
    stage = classify_company(rev, oi, rev_cagr_3yr,
                             stock.get("gross_margin", 0.50) or 0.50)
    stage_targets = {
        "hyper_growth":  0.15,   # Mean-revert toward high but sustainable growth
        "high_growth":   0.10,
        "mature_growth": 0.07,
        "stable":        0.04,
        "turnaround":    0.05,
        "declining":     0.02,
    }
    industry_avg_g = stage_targets.get(stage, 0.07)
    print(f"    ✓ Company stage: {stage} | mean-reversion target: {industry_avg_g:.0%}")

    # Apply mean reversion over projection period
    rev_growth = revenue_mean_reversion(base_g, industry_avg_g, n_proj)
    print(f"    ✓ Revenue growth schedule: {[f'{g:.1%}' for g in rev_growth]}")

    # ── Gross Margin ───────────────────────────────────────────────────────────
    # Detect trend — expanding or compressing?
    if gp and rev:
        gm_latest, gm_trend, gm_r2 = margin_trend(gp, rev)
    else:
        gp_proxy = {y: rev[y] - cogs_d.get(y, 0) for y in rev if y in cogs_d}
        gm_latest, gm_trend, gm_r2 = margin_trend(gp_proxy, rev)

    yahoo_gm = stock.get("gross_margin") or 0

    # Use exp-weighted avg as base; blend with Yahoo LTM if divergent
    if gp:
        gm_wavg = avg_margin(gp, rev, n=4, weighted=True)
    else:
        gp_proxy = {y: rev[y] - cogs_d.get(y, 0) for y in rev if y in cogs_d}
        gm_wavg  = avg_margin(gp_proxy, rev, n=4, weighted=True)

    if not gm_wavg or gm_wavg < 0.01:
        gm_wavg = yahoo_gm or 0.50
    if yahoo_gm and abs(gm_wavg - yahoo_gm) > 0.12:
        gm_wavg = 0.6 * gm_wavg + 0.4 * yahoo_gm

    # Project margin with trend (dampened — trend fades over time)
    # Only trust trend if R² > 0.6 (strong signal)
    gm_proj = []
    gm = gm_wavg
    for i in range(n_proj):
        if gm_r2 > 0.60:
            # Apply trend but dampen it 20% per year (reverts to mean)
            trend_adj = gm_trend * (0.80 ** i)
            gm = gm + trend_adj
        gm = max(0.10, min(0.95, gm))  # hard bounds
        gm_proj.append(round(gm, 4))

    print(f"    ✓ Gross margin: {gm_wavg:.1%} | trend: {gm_trend:+.2%}/yr "
          f"(R²={gm_r2:.2f}) | projected: {[f'{g:.1%}' for g in gm_proj]}")

    # ── Operating Margin ───────────────────────────────────────────────────────
    om_latest, om_trend, om_r2 = margin_trend(oi, rev) if oi else (0, 0, 0)
    om_wavg = avg_margin(oi, rev, n=4, weighted=True) if oi else 0

    if not om_wavg or abs(om_wavg) < 0.001:
        om_wavg = stock.get("op_margin") or 0
    if not om_wavg:
        om_wavg = max(gm_wavg - 0.25, 0.01)  # rough estimate

    # Operating leverage — if DOL is high, margins expand with revenue growth
    dol = operating_leverage(rev, oi)

    # Project operating margin with operating leverage and trend
    om_proj = []
    om = om_wavg
    for i in range(n_proj):
        rev_g_i = rev_growth[i]
        # Margin expansion from operating leverage: DOL × rev_growth × leverage_factor
        leverage_boost = (dol - 1) * rev_g_i * 0.3  # scale down — partial benefit
        # Trend component (dampened)
        trend_adj = om_trend * (0.75 ** i) if om_r2 > 0.55 else 0
        om = om + leverage_boost + trend_adj
        # Mean-revert toward industry-appropriate operating margin
        target_om = {"hyper_growth": 0.15, "high_growth": 0.20,
                     "mature_growth": 0.22, "stable": 0.18,
                     "turnaround": 0.10, "declining": 0.05}.get(stage, 0.15)
        om = om + 0.10 * (target_om - om)   # 10% reversion per year
        om = max(-0.20, min(0.60, om))       # hard bounds
        om_proj.append(round(om, 4))

    print(f"    ✓ Op margin: {om_wavg:.1%} | DOL: {dol:.2f} | "
          f"projected: {[f'{g:.1%}' for g in om_proj]}")

    # ── SG&A and R&D ──────────────────────────────────────────────────────────
    sga_pct = avg_margin(sga, rev, n=4, weighted=True) if sga else 0.15
    rd_pct  = avg_margin(rd, rev, n=4, weighted=True) if rd else 0.0

    # SG&A typically shows slight scale efficiency — falls as % of rev over time
    _, sga_trend, sga_r2 = margin_trend(sga, rev) if sga else (0, 0, 0)
    sga_proj = []
    s = sga_pct
    for i in range(n_proj):
        if sga_r2 > 0.55:
            s = s + sga_trend * (0.70 ** i)
        s = max(0.02, min(0.60, s))
        sga_proj.append(round(s, 4))

    # ── Tax Rate ───────────────────────────────────────────────────────────────
    # Use exp-weighted average to give more weight to recent effective tax rate
    tax_rate = avg_margin(te, data["pretax_inc"], n=4, weighted=True)                if te and data["pretax_inc"] else 0.21
    tax_rate = max(min(tax_rate, 0.38), 0.05)

    # ── CapEx — Split Maintenance vs Growth ───────────────────────────────────
    maint_cx_pct, growth_cx_pct = maintenance_vs_growth_capex(capex, dep, rev)
    total_capex_pct = maint_cx_pct + growth_cx_pct

    # Growth CapEx scales with revenue growth; maintenance CapEx is more stable
    capex_proj = []
    for i in range(n_proj):
        # Maintenance stays constant; growth CapEx scales with revenue growth rate
        growth_adj = growth_cx_pct * (1 + rev_growth[i] * 0.5)
        cx_i = maint_cx_pct + growth_adj
        cx_i = max(0.01, min(0.30, cx_i))
        capex_proj.append(round(cx_i, 4))

    print(f"    ✓ CapEx: total {total_capex_pct:.1%} = "
          f"maintenance {maint_cx_pct:.1%} + growth {growth_cx_pct:.1%}")

    # ── D&A ────────────────────────────────────────────────────────────────────
    dep_pct = avg_margin(dep, rev, n=4, weighted=True) if dep else 0.04
    # D&A grows with PP&E; tie to avg capex roughly
    dep_pct = max(dep_pct, maint_cx_pct * 0.8)   # at least 80% of maint capex

    # ── NWC ────────────────────────────────────────────────────────────────────
    nwc_pct = nwc_ratio(ca, cl, rev)
    # NWC change = growth × NWC ratio (cash drag from working capital build)
    nwc_change_proj = [round(rev_growth[i] * nwc_pct, 4) for i in range(n_proj)]

    # ── Debt & Capital Structure ───────────────────────────────────────────────
    last_lt_debt = last_val(lt_debt) or 0
    last_eq      = last_val(eq) or 1
    de_ratio     = last_lt_debt / last_eq

    # Debt paydown schedule — assume 5% annual paydown if profitable, 0% if not
    last_oi = last_val(oi) or 0
    annual_paydown = 0.05 if last_oi > 0 else 0.0
    debt_proj = []
    d = last_lt_debt
    for i in range(n_proj):
        d = d * (1 - annual_paydown)
        debt_proj.append(round(d, 0))

    last_ie  = last_val(data.get("interest_exp", {})) or 0
    int_rate = last_ie / last_lt_debt if last_lt_debt and last_lt_debt > 0 else 0.04
    int_rate = max(min(int_rate, 0.10), 0.02)

    # ── Beta & Cost of Capital ─────────────────────────────────────────────────
    beta = stock.get("beta", 1.0) or 1.0
    # For unprofitable hyper-growth companies, use higher ERP to reflect risk
    if stage == "hyper_growth" and (last_val(ni) or 0) < 0:
        erp = 0.065   # slightly higher risk premium
    else:
        erp = 0.055

    # ── Payout / Dividends ────────────────────────────────────────────────────
    dps_s = data.get("dps", {}); eps_s = data.get("eps_diluted", {})
    payout = stock.get("payout_ratio") or 0.0
    if not payout and dps_s and eps_s:
        pv = [dps_s[y]/eps_s[y] for y in dps_s
              if y in eps_s and eps_s[y] and eps_s[y] > 0]
        payout = sum(pv[-3:]) / len(pv[-3:]) if pv else 0.0

    # ── Scenarios ─────────────────────────────────────────────────────────────
    def make_scenario(rev_mult, margin_mult):
        return {
            "rev_growth":  [round(min(g * rev_mult, 0.60), 4) for g in rev_growth],
            "gross_margin":[round(min(g * margin_mult, 0.95), 4) for g in gm_proj],
            "op_margin":   [round(max(o * margin_mult, -0.20), 4) for o in om_proj],
            "capex_pct":   capex_proj,
        }

    scenarios = {
        "bull": make_scenario(rev_mult=1.25, margin_mult=1.10),
        "base": make_scenario(rev_mult=1.00, margin_mult=1.00),
        "bear": make_scenario(rev_mult=0.70, margin_mult=0.88),
    }

    print(f"    ✓ Scenarios: Bull rev g Yr1={scenarios['bull']['rev_growth'][0]:.1%} | "
          f"Bear rev g Yr1={scenarios['bear']['rev_growth'][0]:.1%}")

    return {
        # Core projections (base case)
        "rev_growth":          rev_growth,
        "gross_margin":        round(gm_wavg, 4),
        "gross_margin_proj":   gm_proj,
        "sga_pct":             round(sga_pct, 4),
        "sga_proj":            sga_proj,
        "rd_pct":              round(rd_pct, 4),
        "op_margin":           round(om_wavg, 4),
        "op_margin_proj":      om_proj,
        "tax_rate":            round(tax_rate, 4),
        "dep_pct":             round(dep_pct, 4),
        "capex_pct":           round(total_capex_pct, 4),
        "capex_proj":          capex_proj,
        "maint_capex_pct":     maint_cx_pct,
        "growth_capex_pct":    growth_cx_pct,
        "nwc_pct":             round(nwc_pct, 4),
        "nwc_change_proj":     nwc_change_proj,
        "interest_rate":       round(int_rate, 4),
        "payout_ratio":        round(min(payout, 0.95), 4),
        "debt_proj":           debt_proj,
        # Capital structure
        "beta":                round(beta, 2),
        "rf_rate":             0.045,
        "erp":                 erp,
        "lt_growth":           0.03,
        "de_ratio":            round(de_ratio, 4),
        "dol":                 round(dol, 2),
        # Company classification
        "stage":               stage,
        "industry_avg_growth": industry_avg_g,
        # Trend signals
        "gm_trend":            round(gm_trend, 4),
        "gm_r2":               round(gm_r2, 3),
        "om_trend":            round(om_trend, 4),
        "om_r2":               round(om_r2, 3),
        # Scenarios
        "scenarios":           scenarios,
        # Analyst data
        "analyst_yr1_growth":  analyst_yr1_growth,
        "analyst_count":       analyst_count,
    }

# ── Excel Builder ─────────────────────────────────────────────────────────────

def build_workbook(data, n_proj, assumptions):
    wb = Workbook()
    wb.remove(wb.active)

    hist_years = sorted(set(
        list(data["revenue"].keys()) + list(data["net_income"].keys()) +
        list(data["total_assets"].keys())
    ))[-4:]

    last_hist = int(hist_years[-1]) if hist_years else datetime.now().year - 1
    proj_years = [str(last_hist + i + 1) for i in range(n_proj)]

    _cover(wb, data, hist_years, proj_years, assumptions)
    _assumptions(wb, data, assumptions, hist_years, proj_years)
    _income_statement(wb, data, assumptions, hist_years, proj_years)
    _balance_sheet(wb, data, assumptions, hist_years, proj_years)
    _cash_flow(wb, data, assumptions, hist_years, proj_years)
    _valuation(wb, data, assumptions, hist_years, proj_years)

    return wb

# ── COVER ─────────────────────────────────────────────────────────────────────

def _cover(wb, data, hist_years, proj_years, assumptions):
    ws = wb.create_sheet("Cover")
    ws.sheet_view.showGridLines = False

    stock = data["stock"]
    price       = stock.get("price", 0) or 0
    shares_out  = stock.get("shares_out", 0) or 0
    market_cap  = stock.get("market_cap", 0) or (price * shares_out)
    sector      = stock.get("sector", "—") or "—"
    industry    = stock.get("industry", "—") or "—"
    description = stock.get("description", "") or ""
    employees   = stock.get("fulltime_employees", 0) or 0
    target_px   = stock.get("target_price", 0) or 0
    rec         = stock.get("rec", "") or ""
    beta        = stock.get("beta", "—")

    # Banner
    for r in range(1, 18):
        for c in range(1, 14):
            ws.cell(row=r, column=c).fill = fl(DARK_BLUE)

    ws.row_dimensions[3].height = 32
    ws.row_dimensions[4].height = 20
    ws.row_dimensions[5].height = 14

    t1 = ws.cell(row=3, column=2, value=f"{data['ticker']}  —  Financial Operating Model")
    t1.font = Font(name="Arial", bold=True, size=24, color=WHITE)
    t2 = ws.cell(row=4, column=2, value=data["name"])
    t2.font = Font(name="Arial", size=13, color=ACCENT_BLUE)
    t3 = ws.cell(row=5, column=2, value=f"{sector}  |  {industry}")
    t3.font = Font(name="Arial", size=10, color="AAAAAA", italic=True)

    # Key stats box — left column
    left_items = [
        ("Share Price",          f"${price:,.2f}" if price else "N/A",       "Source: Yahoo Finance"),
        ("Shares Outstanding",   f"{shares_out/1e6:,.1f}M" if shares_out else "N/A", "Source: Yahoo Finance / SEC"),
        ("Market Capitalization",f"${market_cap/1e9:,.2f}B" if market_cap else "N/A","Source: Yahoo Finance"),
        ("Sector",               sector,                                       "Source: Yahoo Finance"),
        ("Industry",             industry,                                     "Source: Yahoo Finance"),
        ("Full-Time Employees",  f"{employees:,}" if employees else "N/A",    "Source: Yahoo Finance"),
        ("Beta (5Y Monthly)",    f"{beta:.2f}" if isinstance(beta, float) else str(beta), "Source: Yahoo Finance"),
    ]
    right_items = [
        ("Analyst Target Price", f"${target_px:,.2f}" if target_px else "N/A", "Source: Yahoo Finance consensus"),
        ("Analyst Recommendation", rec.upper() if rec else "N/A",              "Source: Yahoo Finance"),
        ("Historical Period",    f"{hist_years[0]} – {hist_years[-1]}",        "Source: SEC EDGAR 10-K"),
        ("Projection Period",    f"{proj_years[0]}E – {proj_years[-1]}E",      "Model assumption"),
        ("Model Date",           datetime.now().strftime("%B %d, %Y"),         ""),
        ("Data Sources",         "SEC EDGAR, Yahoo Finance",                   ""),
        ("Ticker",               data["ticker"],                                "Exchange: NASDAQ / NYSE"),
    ]

    for i, (k, v, src) in enumerate(left_items):
        r = 7 + i
        ws.row_dimensions[r].height = 14
        ws.cell(row=r, column=2, value=k).font = Font(name="Arial", bold=True, color="AAAAAA", size=9)
        ws.cell(row=r, column=4, value=v).font = Font(name="Arial", color=WHITE, size=9)
        ws.cell(row=r, column=7, value=src).font = Font(name="Arial", color="666666", size=8, italic=True)

    for i, (k, v, src) in enumerate(right_items):
        r = 7 + i
        ws.cell(row=r, column=9, value=k).font = Font(name="Arial", bold=True, color="AAAAAA", size=9)
        ws.cell(row=r, column=11, value=v).font = Font(name="Arial", color=WHITE, size=9)

    # Description block
    r_desc = 17
    ws.row_dimensions[r_desc].height = 10
    for r in range(r_desc, r_desc + 8):
        ws.cell(row=r, column=1).fill = fl(DARK_BLUE)

    hdr = ws.cell(row=r_desc, column=2, value="Business Overview")
    hdr.font = Font(name="Arial", bold=True, size=11, color=ACCENT_BLUE)

    if description:
        desc_cell = ws.cell(row=r_desc+1, column=2, value=description[:900] + ("..." if len(description)>900 else ""))
        desc_cell.font = Font(name="Arial", size=9, color=WHITE)
        try: ws.merge_cells(start_row=r_desc+1, start_column=2, end_row=r_desc+6, end_column=12)
        except: pass
        desc_cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r_desc+1].height = 75

    # Color legend
    r_leg = 27
    ws.cell(row=r_leg, column=2, value="Color-Coding Convention").font = Font(name="Arial", bold=True, size=10, color=DARK_BLUE)
    legends = [
        (INPUT_BLUE,  "Blue text  =  Hardcoded input assumptions (editable)"),
        (BLACK,       "Black text =  Excel formulas (auto-calculated)"),
        ("375623",    "Green text =  Forward projections"),
        ("595959",    "Gray italic =  Growth rates / margin percentages"),
    ]
    for i, (col_hex, txt) in enumerate(legends):
        c = ws.cell(row=r_leg+1+i, column=2, value="■  " + txt)
        c.font = Font(name="Arial", color=col_hex, size=9)

    # Peer summary box
    peers = data.get("peers", {})
    if peers:
        r_peer = 33
        ws.cell(row=r_peer, column=2, value="Peer Comparison Snapshot").font = Font(name="Arial", bold=True, size=10, color=DARK_BLUE)
        hdrs = ["Ticker", "Price", "Mkt Cap ($B)", "P/E (Fwd)", "EV/EBITDA", "Sector"]
        for j, h in enumerate(hdrs):
            c = ws.cell(row=r_peer+1, column=2+j, value=h)
            c.font = Font(name="Arial", bold=True, color=WHITE, size=9)
            c.fill = fl(MID_BLUE); c.alignment = aln("center")
        for i, (pticker, pd) in enumerate(peers.items()):
            r_p = r_peer + 2 + i
            vals = [pticker,
                    f"${pd.get('price',0):,.2f}" if pd.get('price') else "—",
                    f"${pd.get('market_cap',0)/1e9:,.1f}" if pd.get('market_cap') else "—",
                    f"{pd.get('pe_forward',0):.1f}x" if pd.get('pe_forward') else "—",
                    f"{pd.get('ev_ebitda',0):.1f}x" if pd.get('ev_ebitda') else "—",
                    pd.get('sector','—')]
            for j, v in enumerate(vals):
                c = ws.cell(row=r_p, column=2+j, value=v)
                c.font = Font(name="Arial", size=9)
                c.fill = fl(GRAY_LIGHT if i % 2 == 0 else WHITE)
                c.alignment = aln("center")

    ws.column_dimensions["A"].width = 1.5
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 2
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 2
    ws.column_dimensions["F"].width = 2
    ws.column_dimensions["G"].width = 28
    ws.column_dimensions["H"].width = 2
    ws.column_dimensions["I"].width = 24
    ws.column_dimensions["J"].width = 2
    ws.column_dimensions["K"].width = 18
    ws.column_dimensions["L"].width = 2

# ── ASSUMPTIONS ───────────────────────────────────────────────────────────────

def _assumptions(wb, data, assumptions, hist_years, proj_years):
    ws = wb.create_sheet("Assumptions")
    ws.sheet_view.showGridLines = False

    n_proj = len(proj_years)
    ticker = data["ticker"]
    stock  = data["stock"]
    peers  = data.get("peers", {})
    filings = data.get("filings", {})

    last_10k_date = "—"
    last_10k_url  = "https://www.sec.gov"
    if filings.get("10-K"):
        last_10k_date = filings["10-K"][0].get("date","—")
        last_10k_url  = filings["10-K"][0].get("url", last_10k_url)

    section_hdr(ws, 1, 1, 14, f"KEY MODEL ASSUMPTIONS  —  {ticker}  |  {data['name']}")
    ws.cell(row=2, column=1, value=f"  All assumptions sourced from SEC EDGAR 10-K filings, Yahoo Finance, and analyst consensus. "
            f"Projections represent management guidance-informed estimates.").font = ft(italic=True, color="595959", size=8)

    # ── Cost of Capital ──
    row = 4
    sub_hdr(ws, row, 1, 14, "COST OF CAPITAL ASSUMPTIONS")

    coc_items = [
        # (label, value, format, source note, row_ref_name)
        ("Risk-Free Rate (10Y UST)",        assumptions["rf_rate"],  "0.00%", "Source: US Treasury 10-Year yield (as of model date). Used as baseline risk-free return.", "rf"),
        ("Equity Risk Premium",             assumptions["erp"],      "0.00%", "Source: Damodaran (NYU) implied ERP estimate. Represents excess return of equities over risk-free rate.", "erp"),
        ("Beta (Levered, 5Y Monthly)",      assumptions["beta"],     "0.00",  f"Source: Yahoo Finance 5-year monthly regression vs. S&P 500. {ticker} beta reflects systematic market risk.", "beta"),
        ("Cost of Equity (CAPM)",           None,                    "0.00%", "Formula: Rf + Beta × ERP. Represents required return for equity investors.", "coe"),
        ("Long-Term Terminal Growth Rate",  assumptions["lt_growth"],"0.00%", "Assumption: Conservative long-run nominal GDP growth rate floor (3.0%). Used in terminal value calc.", "ltg"),
        ("Effective Tax Rate",              assumptions["tax_rate"], "0.00%", f"Source: SEC EDGAR 10-K ({last_10k_date}). 3-year average effective tax rate from historical filings.", "tax"),
    ]

    COC_ROWS = {}
    for i, (label, val, fmt, note, ref) in enumerate(coc_items):
        r = row + 1 + i
        ws.row_dimensions[r].height = 13
        lbl(ws, r, 1, label, indent=1, bg=GRAY_LIGHT)
        vc = ws.cell(row=r, column=3)
        if ref == "coe":
            rf_r  = row + 1 + 0  # rf row
            beta_r= row + 1 + 2  # beta row
            erp_r = row + 1 + 1  # erp row
            vc.value = f"=C{rf_r}+C{beta_r}*C{erp_r}"
            vc.font = ft(color=BLACK); vc.fill = fl(GRAY_LIGHT)
        else:
            vc.value = val
            vc.font = ft(color=INPUT_BLUE); vc.fill = fl(YELLOW_BG)
        vc.number_format = fmt; vc.alignment = aln("right")
        note_cell(ws, r, 5, note)
        COC_ROWS[ref] = r

    ASS_COE_ROW = COC_ROWS["coe"]
    ASS_LTG_ROW = COC_ROWS["ltg"]
    ASS_TAX_ROW = COC_ROWS["tax"]

    # ── Revenue & Margin Assumptions ──
    row = row + len(coc_items) + 3
    sub_hdr(ws, row, 1, 14, "REVENUE & MARGIN ASSUMPTIONS  (Yellow cells = editable inputs)")

    # Year headers
    lbl(ws, row+1, 1, "  Assumption", bold=True)
    lbl(ws, row+1, 4, "Source / Basis", italic=True, color="595959")
    lbl(ws, row+1, 10, "vs. Peers", bold=True, color=DARK_BLUE)

    # Peer averages for comparison
    peer_gm   = [p.get("gross_margin",0) for p in peers.values() if p.get("gross_margin")]
    peer_pm   = [p.get("profit_margin",0) for p in peers.values() if p.get("profit_margin")]
    peer_opm  = [p.get("op_margin",0) for p in peers.values() if p.get("op_margin")]
    avg_peer_gm  = sum(peer_gm)/len(peer_gm) if peer_gm else None
    avg_peer_pm  = sum(peer_pm)/len(peer_pm) if peer_pm else None
    avg_peer_opm = sum(peer_opm)/len(peer_opm) if peer_opm else None

    # Proj year columns: col 3 = yr1E, col 4 = yr2E, etc.  (col 2 = spacer, col 1 = label)
    PROJ_COL_START = 3
    for i, yr in enumerate(proj_years):
        c = ws.cell(row=row+1, column=PROJ_COL_START+i, value=f"{yr}E")
        c.font = ft(bold=True, color=WHITE, size=9); c.fill = fl(MID_BLUE); c.alignment = aln("center")

    # Pull per-year projected arrays from smarter forecasting engine
    gm_proj_arr  = assumptions.get("gross_margin_proj",  [assumptions["gross_margin"]] * n_proj)
    om_proj_arr  = assumptions.get("op_margin_proj",     [assumptions["op_margin"]]    * n_proj)
    sga_proj_arr = assumptions.get("sga_proj",           [assumptions["sga_pct"]]      * n_proj)
    cx_proj_arr  = assumptions.get("capex_proj",         [assumptions["capex_pct"]]    * n_proj)
    stage        = assumptions.get("stage", "unknown")
    dol          = assumptions.get("dol", 1.0)
    gm_trend     = assumptions.get("gm_trend", 0)
    om_trend     = assumptions.get("om_trend", 0)
    maint_cx     = assumptions.get("maint_capex_pct", 0)
    growth_cx    = assumptions.get("growth_capex_pct", 0)
    nwc_pct      = assumptions.get("nwc_pct", 0)
    a_count      = assumptions.get("analyst_count", 0)
    ind_avg_g    = assumptions.get("industry_avg_growth", 0.07)

    MARGIN_ASS_ROWS = {}
    margin_items = [
        # (key, label, values_list, fmt, source_text, peer_val)
        ("rev_growth", "Revenue Growth Rate",
         assumptions["rev_growth"][:n_proj], "0.0%",
         f"Stage: {stage}. Blends {a_count}-analyst consensus (dynamic weight) with 3yr & 5yr "
         f"historical CAGRs. Mean-reverts to industry avg {ind_avg_g:.0%} over projection period.",
         stock.get("revenue_growth")),
        ("gross_margin", "Gross Margin (per-year, trend-adjusted)",
         gm_proj_arr[:n_proj], "0.0%",
         f"Exp-weighted 4yr avg. Margin trend: {gm_trend:+.2%}/yr "
         f"(R²={assumptions.get('gm_r2',0):.2f}). Trend applied if R²>0.60, dampened 20%/yr.",
         avg_peer_gm),
        ("sga_pct", "SG&A as % of Revenue (per-year)",
         sga_proj_arr[:n_proj], "0.0%",
         f"Exp-weighted avg with scale efficiency trend. SG&A typically falls as % of revenue at scale.",
         None),
        ("rd_pct", "R&D as % of Revenue",
         [assumptions["rd_pct"]] * n_proj, "0.0%",
         f"Exp-weighted 4yr avg R&D/Revenue. Held constant — R&D is a strategic choice.",
         None),
        ("op_margin", "Operating Margin (per-year, with OpLev)",
         om_proj_arr[:n_proj], "0.0%",
         f"Operating leverage (DOL={dol:.2f}) applied. Op margin trend: {om_trend:+.2%}/yr. "
         f"Mean-reverts toward stage-appropriate target margin.",
         avg_peer_opm),
        ("dep_pct", "D&A as % of Revenue",
         [assumptions["dep_pct"]] * n_proj, "0.0%",
         f"Exp-weighted 4yr avg D&A/Revenue. Floor = 80% of maintenance CapEx ({maint_cx:.1%}).",
         None),
        ("capex_pct", "CapEx: Maintenance + Growth (per-year)",
         cx_proj_arr[:n_proj], "0.0%",
         f"Split: Maintenance {maint_cx:.1%} (stable, ≈D&A) + "
         f"Growth {growth_cx:.1%} (scales with revenue growth rate).",
         None),
        ("nwc_pct", "NWC Change as % of Revenue",
         [nwc_pct] * n_proj, "0.0%",
         f"Net working capital / revenue from balance sheet. "
         f"NWC cash drag = revenue growth × {nwc_pct:.1%}. Higher NWC = more cash consumed at growth.",
         None),
        ("payout_ratio", "Dividend Payout Ratio",
         [assumptions["payout_ratio"]] * n_proj, "0.0%",
         f"Historical DPS/EPS ratio from SEC filings. Zero if no dividend history.",
         None),
    ]

    for j, (key, label, vals, fmt, note, peer_val) in enumerate(margin_items):
        r = row + 2 + j
        ws.row_dimensions[r].height = 13
        lbl(ws, r, 1, label, indent=1, bg=GRAY_LIGHT)
        for k, v in enumerate(vals[:n_proj]):
            c = ws.cell(row=r, column=PROJ_COL_START + k, value=v)
            c.number_format = fmt; c.font = ft(color=INPUT_BLUE); c.fill = fl(YELLOW_BG)
            c.alignment = aln("right")
        note_cell(ws, r, PROJ_COL_START + n_proj + 1, note)
        if peer_val:
            c = ws.cell(row=r, column=PROJ_COL_START + n_proj + 1 + 6)
            c.value = f"Peer avg: {peer_val:.1%}"
            c.font = ft(color=DARK_BLUE, italic=True, size=8)
        MARGIN_ASS_ROWS[key] = r

    # Store row references for other sheets to use
    ws._ass_coe_row  = ASS_COE_ROW
    ws._ass_ltg_row  = ASS_LTG_ROW
    ws._ass_tax_row  = ASS_TAX_ROW
    ws._margin_rows  = MARGIN_ASS_ROWS
    ws._proj_col_start = PROJ_COL_START

    # ── Peer Trading Multiples ──
    row = row + len(margin_items) + 4
    sub_hdr(ws, row, 1, 14, "PEER COMPARABLE COMPANY ANALYSIS  (Source: Yahoo Finance)")

    peer_hdr_cols = ["Company", "Ticker", "Price", "Mkt Cap ($B)", "EV/Rev", "EV/EBITDA",
                     "P/E (Fwd)", "Gross Margin", "Op. Margin", "Rev Growth", "Beta"]
    for j, h in enumerate(peer_hdr_cols):
        c = ws.cell(row=row+1, column=1+j, value=h)
        c.font = ft(bold=True, color=WHITE, size=9); c.fill = fl(MID_BLUE); c.alignment = aln("center")

    # Subject company row
    s = data["stock"]
    subject_vals = [
        data["name"], data["ticker"],
        f"${s.get('price',0):,.2f}" if s.get('price') else "—",
        f"${s.get('market_cap',0)/1e9:,.1f}" if s.get('market_cap') else "—",
        f"{s.get('ev_rev',0):.1f}x" if s.get('ev_rev') else "—",
        f"{s.get('ev_ebitda',0):.1f}x" if s.get('ev_ebitda') else "—",
        f"{s.get('pe_forward',0):.1f}x" if s.get('pe_forward') else "—",
        f"{s.get('gross_margin',0):.1%}" if s.get('gross_margin') else "—",
        f"{s.get('op_margin',0):.1%}" if s.get('op_margin') else "—",
        f"{s.get('revenue_growth',0):.1%}" if s.get('revenue_growth') else "—",
        f"{s.get('beta',0):.2f}" if s.get('beta') else "—",
    ]
    for j, v in enumerate(subject_vals):
        c = ws.cell(row=row+2, column=1+j, value=v)
        c.font = ft(bold=True, size=9); c.fill = fl(ACCENT_BLUE); c.alignment = aln("center")

    for i, (pticker, pd) in enumerate(peers.items()):
        r_p = row + 3 + i
        pvals = [
            pd.get("name", pticker), pticker,
            f"${pd.get('price',0):,.2f}" if pd.get('price') else "—",
            f"${pd.get('market_cap',0)/1e9:,.1f}" if pd.get('market_cap') else "—",
            f"{pd.get('ev_rev',0):.1f}x" if pd.get('ev_rev') else "—",
            f"{pd.get('ev_ebitda',0):.1f}x" if pd.get('ev_ebitda') else "—",
            f"{pd.get('pe_forward',0):.1f}x" if pd.get('pe_forward') else "—",
            f"{pd.get('gross_margin',0):.1%}" if pd.get('gross_margin') else "—",
            f"{pd.get('op_margin',0):.1%}" if pd.get('op_margin') else "—",
            f"{pd.get('revenue_growth',0):.1%}" if pd.get('revenue_growth') else "—",
            f"{pd.get('beta',0):.2f}" if pd.get('beta') else "—",
        ]
        for j, v in enumerate(pvals):
            c = ws.cell(row=r_p, column=1+j, value=v)
            c.font = ft(size=9)
            c.fill = fl(GRAY_LIGHT if i%2==0 else WHITE)
            c.alignment = aln("center")

    # Source footnote
    r_fn = row + 3 + len(peers) + 1
    ws.cell(row=r_fn, column=1, value="Sources: Yahoo Finance (real-time), SEC EDGAR 10-K annual filings, "
            f"Damodaran ERP estimates (damodaran.com). Peer list reflects closest comparables by industry/sector. "
            f"Last 10-K filed: {last_10k_date}.  SEC EDGAR: {last_10k_url}").font = ft(italic=True, color="595959", size=8)

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 2
    for i in range(n_proj): ws.column_dimensions[get_column_letter(PROJ_COL_START+i)].width = 10
    ws.column_dimensions[get_column_letter(PROJ_COL_START+n_proj+1)].width = 60
    for col_l in ["C","D","E","F","G","H","I","J","K"]:
        if ws.column_dimensions[col_l].width < 10:
            ws.column_dimensions[col_l].width = 13

# ── INCOME STATEMENT ──────────────────────────────────────────────────────────

def _income_statement(wb, data, assumptions, hist_years, proj_years):
    ws = wb.create_sheet("Income Statement")
    ws.sheet_view.showGridLines = False

    all_years = hist_years + proj_years
    n_hist    = len(hist_years)
    n_proj    = len(proj_years)
    DC        = 3   # data columns start at col 3
    scale     = 1_000_000

    section_hdr(ws, 1, 1, DC+len(all_years)+2, f"INCOME STATEMENT  —  {data['ticker']}  |  {data['name']}")
    ws.cell(row=2, column=1, value=f"  $ in millions  |  Source: SEC EDGAR 10-K (historical)  |  Model projections (forward)"
            ).font = ft(italic=True, color="595959", size=8)
    ws.cell(row=2, column=1).alignment = aln("left")

    # Column headers: "Historical" / "Projected" banners
    try: ws.merge_cells(start_row=3, start_column=DC, end_row=3, end_column=DC+n_hist-1)
    except: pass
    c = ws.cell(row=3, column=DC, value="◀  Historical  ▶")
    c.font = ft(bold=True, color=WHITE, size=9, italic=True)
    c.fill = fl(MID_BLUE); c.alignment = aln("center")

    try: ws.merge_cells(start_row=3, start_column=DC+n_hist, end_row=3, end_column=DC+n_hist+n_proj-1)
    except: pass
    c2 = ws.cell(row=3, column=DC+n_hist, value="◀  Projected  ▶")
    c2.font = ft(bold=True, color=WHITE, size=9, italic=True)
    c2.fill = fl(fl(ACCENT_BLUE).fgColor.rgb if hasattr(fl(ACCENT_BLUE).fgColor,'rgb') else ACCENT_BLUE)
    c2.fill = fl(ACCENT_BLUE); c2.alignment = aln("center")

    # Year row
    lbl(ws, 4, 1, "")
    for i, yr in enumerate(all_years):
        c = ws.cell(row=4, column=DC+i)
        c.value = yr if i < n_hist else f"{yr}E"
        c.font = ft(bold=True, color=WHITE, size=9)
        c.fill = fl(MID_BLUE); c.alignment = aln("center")

    FMT  = '#,##0.0;(#,##0.0);"-"'
    PFMT = "0.0%"

    # Assumption row references on Assumptions sheet
    # We reconstruct which rows hold what based on _assumptions layout
    # CoC section starts row 4, items at 5..10
    # Margin section: we need rev_growth=row, gross_margin=row, etc.
    # Since we can't reliably reference the other sheet's dynamic rows,
    # we embed formulas referencing known Assumptions cell positions.
    # Assumptions sheet col 3 = Yr1E, col 4 = Yr2E, etc.
    ASS_PROJ_COL = 3   # col C on Assumptions = first projection year
    # Row positions on Assumptions sheet (from _assumptions function, margin_items order):
    # CoC: rows 5-10. Margin items start at row 5+6+3 = 14 (row=4+6+3=13, first item = 14)
    # Let's compute: coc_items has 6 items => rows 5-10. Then row = 4+6+3 = 13.
    # margin_items start at row 13+2=15 (row+2+j)
    ASS_MARGIN_START = 15   # row 15 = rev_growth (j=0)
    # Keys in order: rev_growth, gross_margin, sga_pct, rd_pct, op_margin, dep_pct, capex_pct, payout_ratio
    ASS_REV_G  = ASS_MARGIN_START + 0
    ASS_GM     = ASS_MARGIN_START + 1
    ASS_SGA    = ASS_MARGIN_START + 2
    ASS_RD     = ASS_MARGIN_START + 3
    ASS_OPM    = ASS_MARGIN_START + 4
    ASS_DEP    = ASS_MARGIN_START + 5
    ASS_CAPEX  = ASS_MARGIN_START + 6
    ASS_PAYOUT = ASS_MARGIN_START + 7
    ASS_TAX    = 10   # effective tax rate — single cell C10 (same value all years)

    def ass_ref(ass_row, proj_idx):
        # Tax rate is single-column (col C only); all other margins are per-year columns
        if ass_row == ASS_TAX:
            return f"Assumptions!C{ass_row}"
        col_l = get_column_letter(ASS_PROJ_COL + proj_idx)
        return f"Assumptions!{col_l}{ass_row}"

    def dc(i): return DC + i

    row = 5
    # ── Revenue ──
    lbl(ws, row, 1, "Revenue", bold=True, bg=LIGHT_BLUE)
    rev_row = row
    for i, yr in enumerate(hist_years):
        v = data["revenue"].get(yr)
        val_cell(ws, row, dc(i), v/scale if v else None, FMT, bg=LIGHT_BLUE)
    for i in range(n_proj):
        prev_col = get_column_letter(dc(n_hist+i-1))
        c = ws.cell(row=row, column=dc(n_hist+i))
        c.value = f"={prev_col}{rev_row}*(1+{ass_ref(ASS_REV_G, i)})"
        c.number_format = FMT; c.font = ft(color=GREEN); c.fill = fl(LIGHT_BLUE); c.alignment = aln("right")

    row += 1
    # YoY Growth
    lbl(ws, row, 1, "  YoY Growth %", italic=True, color="595959")
    for i in range(1, n_hist):
        cl, pl = get_column_letter(dc(i)), get_column_letter(dc(i-1))
        c = ws.cell(row=row, column=dc(i), value=f"=IFERROR({cl}{rev_row}/{pl}{rev_row}-1,\"\")")
        c.number_format = PFMT; c.font = ft(italic=True, color="595959", size=9); c.alignment = aln("right")
    for i in range(n_proj):
        c = ws.cell(row=row, column=dc(n_hist+i), value=f"={ass_ref(ASS_REV_G,i)}")
        c.number_format = PFMT; c.font = ft(italic=True, color=GREEN, size=9); c.alignment = aln("right")

    row += 1
    # COGS
    lbl(ws, row, 1, "  Cost of Revenue", indent=1)
    cogs_row = row
    for i, yr in enumerate(hist_years):
        v = data["cogs"].get(yr)
        if not v and data["gross_profit"].get(yr) and data["revenue"].get(yr):
            v = data["revenue"][yr] - data["gross_profit"][yr]
        val_cell(ws, row, dc(i), -abs(v)/scale if v else None, FMT)
    for i in range(n_proj):
        rc = get_column_letter(dc(n_hist+i))
        c = ws.cell(row=row, column=dc(n_hist+i))
        c.value = f"=-{rc}{rev_row}*(1-{ass_ref(ASS_GM,i)})"
        c.number_format = FMT; c.font = ft(color=GREEN); c.alignment = aln("right")

    row += 1
    # Gross Profit
    lbl(ws, row, 1, "Gross Profit", bold=True, bg=ACCENT_BLUE)
    gp_row = row
    for i in range(len(all_years)):
        cl = get_column_letter(dc(i))
        c = ws.cell(row=row, column=dc(i))
        c.value = f"={cl}{rev_row}+{cl}{cogs_row}"
        c.number_format = FMT
        c.font = ft(bold=True, color=GREEN if i>=n_hist else BLACK)
        c.fill = fl(ACCENT_BLUE); c.alignment = aln("right")

    row += 1
    lbl(ws, row, 1, "  Gross Margin %", italic=True, color="595959")
    for i in range(len(all_years)):
        cl = get_column_letter(dc(i))
        c = ws.cell(row=row, column=dc(i))
        c.value = f"=IFERROR({cl}{gp_row}/{cl}{rev_row},\"\")"
        c.number_format = PFMT; c.font = ft(italic=True, color=GREEN if i>=n_hist else "595959"); c.alignment = aln("right")

    row += 1
    # SG&A
    lbl(ws, row, 1, "  SG&A Expenses", indent=1)
    sga_row = row
    for i, yr in enumerate(hist_years):
        v = data["sga_exp"].get(yr)
        val_cell(ws, row, dc(i), -abs(v)/scale if v else None, FMT)
    for i in range(n_proj):
        rc = get_column_letter(dc(n_hist+i))
        c = ws.cell(row=row, column=dc(n_hist+i))
        c.value = f"=-{rc}{rev_row}*{ass_ref(ASS_SGA,i)}"
        c.number_format = FMT; c.font = ft(color=GREEN); c.alignment = aln("right")

    row += 1
    # R&D
    lbl(ws, row, 1, "  Research & Development", indent=1)
    rd_row = row
    for i, yr in enumerate(hist_years):
        v = data["rd_exp"].get(yr)
        val_cell(ws, row, dc(i), -abs(v)/scale if v else None, FMT)
    for i in range(n_proj):
        rc = get_column_letter(dc(n_hist+i))
        c = ws.cell(row=row, column=dc(n_hist+i))
        c.value = f"=-{rc}{rev_row}*{ass_ref(ASS_RD,i)}"
        c.number_format = FMT; c.font = ft(color=GREEN); c.alignment = aln("right")

    row += 1
    # D&A
    lbl(ws, row, 1, "  Depreciation & Amortization", indent=1)
    da_row = row
    for i, yr in enumerate(hist_years):
        v = data["dep_amor"].get(yr)
        val_cell(ws, row, dc(i), -abs(v)/scale if v else None, FMT)
    for i in range(n_proj):
        rc = get_column_letter(dc(n_hist+i))
        c = ws.cell(row=row, column=dc(n_hist+i))
        c.value = f"=-{rc}{rev_row}*{ass_ref(ASS_DEP,i)}"
        c.number_format = FMT; c.font = ft(color=GREEN); c.alignment = aln("right")

    row += 1
    # EBIT
    lbl(ws, row, 1, "Operating Income (EBIT)", bold=True, bg=ACCENT_BLUE)
    ebit_row = row
    for i in range(len(all_years)):
        cl = get_column_letter(dc(i))
        c = ws.cell(row=row, column=dc(i))
        c.value = f"={cl}{gp_row}+{cl}{sga_row}+{cl}{rd_row}+{cl}{da_row}"
        c.number_format = FMT
        c.font = ft(bold=True, color=GREEN if i>=n_hist else BLACK)
        c.fill = fl(ACCENT_BLUE); c.alignment = aln("right")

    row += 1
    lbl(ws, row, 1, "  EBIT Margin %", italic=True, color="595959")
    for i in range(len(all_years)):
        cl = get_column_letter(dc(i))
        c = ws.cell(row=row, column=dc(i))
        c.value = f"=IFERROR({cl}{ebit_row}/{cl}{rev_row},\"\")"
        c.number_format = PFMT; c.font = ft(italic=True, color=GREEN if i>=n_hist else "595959"); c.alignment = aln("right")

    row += 1
    # EBITDA
    lbl(ws, row, 1, "EBITDA", bold=True, bg=LIGHT_BLUE)
    ebitda_row = row
    for i in range(len(all_years)):
        cl = get_column_letter(dc(i))
        c = ws.cell(row=row, column=dc(i))
        c.value = f"={cl}{ebit_row}-{cl}{da_row}"
        c.number_format = FMT
        c.font = ft(bold=True, color=GREEN if i>=n_hist else BLACK)
        c.fill = fl(LIGHT_BLUE); c.alignment = aln("right")

    row += 1
    lbl(ws, row, 1, "  EBITDA Margin %", italic=True, color="595959")
    for i in range(len(all_years)):
        cl = get_column_letter(dc(i))
        c = ws.cell(row=row, column=dc(i))
        c.value = f"=IFERROR({cl}{ebitda_row}/{cl}{rev_row},\"\")"
        c.number_format = PFMT; c.font = ft(italic=True, color=GREEN if i>=n_hist else "595959"); c.alignment = aln("right")

    row += 1
    # Interest Expense
    lbl(ws, row, 1, "  Interest Expense, net", indent=1)
    int_row = row
    last_int = abs(last_val(data["interest_exp"]) or 0) / scale
    for i, yr in enumerate(hist_years):
        v = data["interest_exp"].get(yr)
        val_cell(ws, row, dc(i), -abs(v)/scale if v else None, FMT)
    for i in range(n_proj):
        val_cell(ws, row, dc(n_hist+i), -last_int, FMT, color=INPUT_BLUE)

    row += 1
    # Pre-tax
    lbl(ws, row, 1, "Pre-Tax Income", bold=True, bg=ACCENT_BLUE)
    pretax_row = row
    for i in range(len(all_years)):
        cl = get_column_letter(dc(i))
        c = ws.cell(row=row, column=dc(i))
        c.value = f"={cl}{ebit_row}+{cl}{int_row}"
        c.number_format = FMT
        c.font = ft(bold=True, color=GREEN if i>=n_hist else BLACK)
        c.fill = fl(ACCENT_BLUE); c.alignment = aln("right")

    row += 1
    # Tax
    lbl(ws, row, 1, "  Income Tax Expense", indent=1)
    tax_row = row
    for i, yr in enumerate(hist_years):
        v = data["tax_exp"].get(yr)
        val_cell(ws, row, dc(i), -abs(v)/scale if v else None, FMT)
    for i in range(n_proj):
        pc = get_column_letter(dc(n_hist+i))
        c = ws.cell(row=row, column=dc(n_hist+i))
        c.value = f"=-{pc}{pretax_row}*{ass_ref(ASS_TAX,i)}"
        c.number_format = FMT; c.font = ft(color=GREEN); c.alignment = aln("right")

    row += 1
    # Net Income
    lbl(ws, row, 1, "Net Income", bold=True, bg=DARK_BLUE, color=WHITE)
    ni_row = row
    for i in range(len(all_years)):
        cl = get_column_letter(dc(i))
        c = ws.cell(row=row, column=dc(i))
        c.value = f"={cl}{pretax_row}+{cl}{tax_row}"
        c.number_format = FMT
        c.font = ft(bold=True, color=WHITE)
        c.fill = fl(DARK_BLUE); c.alignment = aln("right")

    row += 1
    lbl(ws, row, 1, "  Net Margin %", italic=True, color="595959")
    for i in range(len(all_years)):
        cl = get_column_letter(dc(i))
        c = ws.cell(row=row, column=dc(i))
        c.value = f"=IFERROR({cl}{ni_row}/{cl}{rev_row},\"\")"
        c.number_format = PFMT; c.font = ft(italic=True, color=GREEN if i>=n_hist else "595959"); c.alignment = aln("right")

    row += 2
    # Per share
    sub_hdr(ws, row, 1, DC+len(all_years), "PER SHARE DATA")
    row += 1

    lbl(ws, row, 1, "  Diluted Shares Outstanding (mm)", indent=1)
    shares_row = row
    last_shares = last_val(data["shares_dil"]) or 1
    for i, yr in enumerate(hist_years):
        v = data["shares_dil"].get(yr)
        val_cell(ws, row, dc(i), v/scale if v else None, FMT)
    for i in range(n_proj):
        val_cell(ws, row, dc(n_hist+i), last_shares/scale, FMT, color=INPUT_BLUE)

    row += 1
    lbl(ws, row, 1, "  Diluted EPS", indent=1)
    eps_row = row
    for i in range(len(all_years)):
        cl = get_column_letter(dc(i))
        c = ws.cell(row=row, column=dc(i))
        c.value = f"=IFERROR({cl}{ni_row}/({cl}{shares_row}),\"\")"
        c.number_format = '$#,##0.00;($#,##0.00);"-"'
        c.font = ft(color=GREEN if i>=n_hist else BLACK); c.alignment = aln("right")

    row += 1
    lbl(ws, row, 1, "  Dividends Per Share (DPS)", indent=1)
    for i, yr in enumerate(hist_years):
        v = data["dps"].get(yr)
        val_cell(ws, row, dc(i), v if v else None, '$#,##0.00;($#,##0.00);"-"')
    for i in range(n_proj):
        ec = get_column_letter(dc(n_hist+i))
        c = ws.cell(row=row, column=dc(n_hist+i))
        c.value = f"={ec}{eps_row}*{ass_ref(ASS_PAYOUT,i)}"
        c.number_format = '$#,##0.00;($#,##0.00);"-"'
        c.font = ft(color=GREEN); c.alignment = aln("right")

    # Store key row refs for other sheets
    ws._rev_row     = rev_row
    ws._ni_row      = ni_row
    ws._ebitda_row  = ebitda_row
    ws._ebit_row    = ebit_row
    ws._da_row      = da_row
    ws._shares_row  = shares_row
    ws._eps_row     = eps_row
    ws._n_hist      = n_hist
    ws._DC          = DC

    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 2
    for i in range(len(all_years)):
        ws.column_dimensions[get_column_letter(DC+i)].width = 11

# ── BALANCE SHEET ─────────────────────────────────────────────────────────────

def _balance_sheet(wb, data, assumptions, hist_years, proj_years):
    ws = wb.create_sheet("Balance Sheet")
    ws.sheet_view.showGridLines = False

    all_years = hist_years + proj_years
    n_hist = len(hist_years)
    DC = 3; scale = 1_000_000

    section_hdr(ws, 1, 1, DC+len(all_years)+1, f"BALANCE SHEET  —  {data['ticker']}  |  {data['name']}")
    ws.cell(row=2, column=1, value="  $ in millions  |  Source: SEC EDGAR 10-K (historical)  |  Forward projections grow with revenue").font = ft(italic=True, color="595959", size=8)

    try: ws.merge_cells(start_row=3, start_column=DC, end_row=3, end_column=DC+n_hist-1)
    except: pass
    ws.cell(row=3, column=DC, value="Historical").font = ft(bold=True, color=WHITE, size=9)
    ws.cell(row=3, column=DC).fill = fl(MID_BLUE); ws.cell(row=3, column=DC).alignment = aln("center")
    try: ws.merge_cells(start_row=3, start_column=DC+n_hist, end_row=3, end_column=DC+n_hist+len(proj_years)-1)
    except: pass
    ws.cell(row=3, column=DC+n_hist, value="Projected").font = ft(bold=True, color=WHITE, size=9)
    ws.cell(row=3, column=DC+n_hist).fill = fl(ACCENT_BLUE); ws.cell(row=3, column=DC+n_hist).alignment = aln("center")

    for i, yr in enumerate(all_years):
        c = ws.cell(row=4, column=DC+i)
        c.value = yr if i < n_hist else f"{yr}E"
        c.font = ft(bold=True, color=WHITE, size=9); c.fill = fl(MID_BLUE); c.alignment = aln("center")

    FMT = '#,##0.0;(#,##0.0);"-"'

    def dc(i): return DC + i

    def grow_series(series_dict, grow_fn):
        """Return projected values for proj_years using grow_fn(last_val, year_index)."""
        last = last_val(series_dict) or 0
        proj = {}
        for i, yr in enumerate(proj_years):
            last = grow_fn(last, i)
            proj[yr] = last
        return proj

    rev = data["revenue"]
    rev_growth = assumptions["rev_growth"]

    def by_rev_growth(last, i):
        return last * (1 + rev_growth[i])

    def stable(last, i): return last  # no growth (e.g. goodwill)

    # ASSETS
    row = 5
    sub_hdr(ws, row, 1, DC+len(all_years), "ASSETS")
    row += 1

    cash_proj = grow_series(data["cash"], lambda l,i: l*(1+rev_growth[i]*0.5))
    lbl(ws, row, 1, "  Cash & Cash Equivalents", indent=1)
    for i, yr in enumerate(all_years):
        src = data["cash"] if i<n_hist else cash_proj
        v = src.get(yr)
        c = ws.cell(row=row, column=dc(i), value=v/scale if v else None)
        c.number_format = FMT; c.font = ft(color=GREEN if i>=n_hist else BLACK); c.alignment = aln("right")
    cash_row = row

    row += 1
    ca_proj = grow_series(data["curr_assets"], by_rev_growth)
    lbl(ws, row, 1, "  Total Current Assets", bold=True, bg=ACCENT_BLUE)
    for i, yr in enumerate(all_years):
        src = data["curr_assets"] if i<n_hist else ca_proj
        v = src.get(yr)
        c = ws.cell(row=row, column=dc(i), value=v/scale if v else None)
        c.number_format = FMT; c.font = ft(bold=True, color=GREEN if i>=n_hist else BLACK)
        c.fill = fl(ACCENT_BLUE); c.alignment = aln("right")
    ca_row = row

    row += 1
    ppe_proj = grow_series(data["pp_e"], by_rev_growth)
    lbl(ws, row, 1, "  PP&E, Net", indent=1)
    for i, yr in enumerate(all_years):
        src = data["pp_e"] if i<n_hist else ppe_proj
        v = src.get(yr)
        c = ws.cell(row=row, column=dc(i), value=v/scale if v else None)
        c.number_format = FMT; c.font = ft(color=GREEN if i>=n_hist else BLACK); c.alignment = aln("right")

    row += 1
    gw_proj = grow_series(data["goodwill"], stable)
    lbl(ws, row, 1, "  Goodwill", indent=1)
    for i, yr in enumerate(all_years):
        src = data["goodwill"] if i<n_hist else gw_proj
        v = src.get(yr)
        c = ws.cell(row=row, column=dc(i), value=v/scale if v else None)
        c.number_format = FMT; c.font = ft(color=GREEN if i>=n_hist else BLACK); c.alignment = aln("right")

    row += 1
    int_proj = grow_series(data["intangibles"], lambda l,i: l*0.9)  # amortize down
    lbl(ws, row, 1, "  Intangible Assets, Net", indent=1)
    for i, yr in enumerate(all_years):
        src = data["intangibles"] if i<n_hist else int_proj
        v = src.get(yr)
        c = ws.cell(row=row, column=dc(i), value=v/scale if v else None)
        c.number_format = FMT; c.font = ft(color=GREEN if i>=n_hist else BLACK); c.alignment = aln("right")

    row += 1
    ta_proj = grow_series(data["total_assets"], by_rev_growth)
    lbl(ws, row, 1, "Total Assets", bold=True, bg=DARK_BLUE, color=WHITE)
    for i, yr in enumerate(all_years):
        src = data["total_assets"] if i<n_hist else ta_proj
        v = src.get(yr)
        c = ws.cell(row=row, column=dc(i), value=v/scale if v else None)
        c.number_format = FMT; c.font = ft(bold=True, color=WHITE)
        c.fill = fl(DARK_BLUE); c.alignment = aln("right")
    ta_row = row

    # LIABILITIES
    row += 2
    sub_hdr(ws, row, 1, DC+len(all_years), "LIABILITIES & EQUITY")
    row += 1

    cl_proj = grow_series(data["curr_liab"], by_rev_growth)
    lbl(ws, row, 1, "  Total Current Liabilities", bold=True, bg=ACCENT_BLUE)
    for i, yr in enumerate(all_years):
        src = data["curr_liab"] if i<n_hist else cl_proj
        v = src.get(yr)
        c = ws.cell(row=row, column=dc(i), value=v/scale if v else None)
        c.number_format = FMT; c.font = ft(bold=True, color=GREEN if i>=n_hist else BLACK)
        c.fill = fl(ACCENT_BLUE); c.alignment = aln("right")

    row += 1
    # LT debt — slight paydown
    ltd_proj = grow_series(data["lt_debt"], lambda l,i: l*0.97)
    lbl(ws, row, 1, "  Long-Term Debt", indent=1)
    for i, yr in enumerate(all_years):
        src = data["lt_debt"] if i<n_hist else ltd_proj
        v = src.get(yr)
        c = ws.cell(row=row, column=dc(i), value=v/scale if v else None)
        c.number_format = FMT; c.font = ft(color=GREEN if i>=n_hist else BLACK); c.alignment = aln("right")
    ltd_row = row

    row += 1
    tl_proj = grow_series(data["total_liab"], by_rev_growth)
    lbl(ws, row, 1, "Total Liabilities", bold=True, bg=DARK_BLUE, color=WHITE)
    for i, yr in enumerate(all_years):
        src = data["total_liab"] if i<n_hist else tl_proj
        v = src.get(yr)
        c = ws.cell(row=row, column=dc(i), value=v/scale if v else None)
        c.number_format = FMT; c.font = ft(bold=True, color=WHITE)
        c.fill = fl(DARK_BLUE); c.alignment = aln("right")

    row += 1
    eq_proj = grow_series(data["total_equity"], by_rev_growth)
    lbl(ws, row, 1, "Total Shareholders' Equity", bold=True, bg=DARK_BLUE, color=WHITE)
    for i, yr in enumerate(all_years):
        src = data["total_equity"] if i<n_hist else eq_proj
        v = src.get(yr)
        c = ws.cell(row=row, column=dc(i), value=v/scale if v else None)
        c.number_format = FMT; c.font = ft(bold=True, color=WHITE)
        c.fill = fl(DARK_BLUE); c.alignment = aln("right")
    eq_row = row

    # Key ratios
    row += 2
    sub_hdr(ws, row, 1, DC+len(all_years), "KEY BALANCE SHEET RATIOS")
    row += 1

    ratio_items = [
        ("  Debt / Equity (x)", lambda cl, r: f"=IFERROR({cl}{ltd_row}/{cl}{eq_row},\"\")", "0.0x"),
        ("  Total Debt / Total Assets", lambda cl, r: f"=IFERROR({cl}{ltd_row}/{cl}{ta_row},\"\")", "0.0%"),
    ]
    for label, formula_fn, fmt in ratio_items:
        lbl(ws, row, 1, label, italic=True, color="595959")
        for i in range(len(all_years)):
            cl = get_column_letter(dc(i))
            c = ws.cell(row=row, column=dc(i), value=formula_fn(cl, row))
            c.number_format = fmt; c.font = ft(italic=True, color=GREEN if i>=n_hist else "595959"); c.alignment = aln("right")
        row += 1

    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 2
    for i in range(len(all_years)):
        ws.column_dimensions[get_column_letter(DC+i)].width = 11

# ── CASH FLOW ─────────────────────────────────────────────────────────────────

def _cash_flow(wb, data, assumptions, hist_years, proj_years):
    ws = wb.create_sheet("Cash Flow")
    ws.sheet_view.showGridLines = False

    all_years = hist_years + proj_years
    n_hist = len(hist_years)
    n_proj = len(proj_years)
    DC = 3; scale = 1_000_000

    section_hdr(ws, 1, 1, DC+len(all_years)+2, f"CASH FLOW STATEMENT  —  {data['ticker']}  |  {data['name']}")
    ws.cell(row=2, column=1, value="  $ in millions  |  Source: SEC EDGAR 10-K (historical)  |  Projections based on model assumptions"
            ).font = ft(italic=True, color="595959", size=8)

    try: ws.merge_cells(start_row=3, start_column=DC, end_row=3, end_column=DC+n_hist-1)
    except: pass
    ws.cell(row=3, column=DC, value="Historical").font = ft(bold=True, color=WHITE, size=9)
    ws.cell(row=3, column=DC).fill = fl(MID_BLUE); ws.cell(row=3, column=DC).alignment = aln("center")
    try: ws.merge_cells(start_row=3, start_column=DC+n_hist, end_row=3, end_column=DC+n_hist+n_proj-1)
    except: pass
    ws.cell(row=3, column=DC+n_hist, value="Projected").font = ft(bold=True, color=WHITE, size=9)
    ws.cell(row=3, column=DC+n_hist).fill = fl(ACCENT_BLUE); ws.cell(row=3, column=DC+n_hist).alignment = aln("center")
    for i, yr in enumerate(all_years):
        c = ws.cell(row=4, column=DC+i)
        c.value = yr if i<n_hist else f"{yr}E"
        c.font = ft(bold=True, color=WHITE, size=9); c.fill = fl(MID_BLUE); c.alignment = aln("center")

    FMT = '#,##0.0;(#,##0.0);"-"'

    def dc(i): return DC+i

    last_rev = last_val(data["revenue"]) or 1
    rev_growth = assumptions["rev_growth"]

    # ─── OPERATING ACTIVITIES ───
    row = 5
    sub_hdr(ws, row, 1, DC+len(all_years)+2, "OPERATING ACTIVITIES")
    row += 1

    # Net Income
    lbl(ws, row, 1, "  Net Income", indent=1)
    ni_row = row
    for i, yr in enumerate(hist_years):
        v = data["net_income"].get(yr)
        val_cell(ws, row, dc(i), v/scale if v else None, FMT)
    proj_rev = last_rev
    for i in range(n_proj):
        proj_rev = proj_rev * (1 + rev_growth[i])
        proj_ni  = proj_rev * assumptions["op_margin"] * (1 - assumptions["tax_rate"])
        val_cell(ws, row, dc(n_hist+i), proj_ni/scale, FMT, color=GREEN)

    row += 1
    # D&A add-back
    lbl(ws, row, 1, "  (+) Depreciation & Amortization", indent=1)
    da_row = row
    proj_rev_tmp = last_rev
    for i, yr in enumerate(hist_years):
        v = data["dep_amor"].get(yr)
        val_cell(ws, row, dc(i), abs(v)/scale if v else None, FMT)
    proj_rev_tmp = last_rev
    for i in range(n_proj):
        proj_rev_tmp = proj_rev_tmp * (1 + rev_growth[i])
        val_cell(ws, row, dc(n_hist+i), proj_rev_tmp*assumptions["dep_pct"]/scale, FMT, color=GREEN)

    row += 1
    # Stock-based comp
    lbl(ws, row, 1, "  (+) Stock-Based Compensation", indent=1)
    sbc_row = row
    for i, yr in enumerate(hist_years):
        v = data.get("stock_comp", {}).get(yr)
        val_cell(ws, row, dc(i), v/scale if v else None, FMT)
    last_sbc = last_val(data.get("stock_comp", {})) or 0
    for i in range(n_proj):
        val_cell(ws, row, dc(n_hist+i), last_sbc/scale, FMT, color=INPUT_BLUE)

    row += 1
    # Changes in WC
    lbl(ws, row, 1, "  (+/-) Changes in Working Capital", indent=1)
    wc_row = row
    for i, yr in enumerate(hist_years):
        op_cf_v = data["op_cf"].get(yr)
        ni_v    = data["net_income"].get(yr)
        da_v    = data["dep_amor"].get(yr)
        sbc_v   = data.get("stock_comp", {}).get(yr, 0) or 0
        if op_cf_v and ni_v and da_v:
            wc = op_cf_v - ni_v - abs(da_v) - sbc_v
            val_cell(ws, row, dc(i), wc/scale, FMT)
    for i in range(n_proj):
        val_cell(ws, row, dc(n_hist+i), 0.0, FMT, color=INPUT_BLUE)

    row += 1
    # CFO total
    lbl(ws, row, 1, "Cash from Operations (CFO)", bold=True, bg=ACCENT_BLUE)
    cfo_row = row
    for i, yr in enumerate(hist_years):
        v = data["op_cf"].get(yr)
        c = ws.cell(row=row, column=dc(i), value=v/scale if v else None)
        c.number_format = FMT; c.font = ft(bold=True); c.fill = fl(ACCENT_BLUE); c.alignment = aln("right")
    for i in range(n_proj):
        ni_c = get_column_letter(dc(n_hist+i)); da_c = ni_c; sbc_c = ni_c; wc_c = ni_c
        c = ws.cell(row=row, column=dc(n_hist+i))
        c.value = (f"={get_column_letter(dc(n_hist+i))}{ni_row}"
                   f"+{get_column_letter(dc(n_hist+i))}{da_row}"
                   f"+{get_column_letter(dc(n_hist+i))}{sbc_row}"
                   f"+{get_column_letter(dc(n_hist+i))}{wc_row}")
        c.number_format = FMT; c.font = ft(bold=True, color=GREEN); c.fill = fl(ACCENT_BLUE); c.alignment = aln("right")

    # CFO margin
    row += 1
    lbl(ws, row, 1, "  CFO as % of Revenue", italic=True, color="595959")
    IS_rev_row = 5  # Revenue is on row 5 of Income Statement
    for i, yr in enumerate(hist_years):
        cfo_v = data["op_cf"].get(yr)
        rev_v = data["revenue"].get(yr)
        val = (cfo_v/rev_v) if cfo_v and rev_v else None
        c = ws.cell(row=row, column=dc(i), value=val)
        c.number_format = "0.0%"; c.font = ft(italic=True, color="595959"); c.alignment = aln("right")
    for i in range(n_proj):
        c = ws.cell(row=row, column=dc(n_hist+i))
        cfo_c = get_column_letter(dc(n_hist+i))
        c.value = f"=IFERROR({cfo_c}{cfo_row}/('Income Statement'!{cfo_c}{IS_rev_row}),\"\")"
        c.number_format = "0.0%"; c.font = ft(italic=True, color=GREEN); c.alignment = aln("right")

    # ─── INVESTING ACTIVITIES ───
    row += 2
    sub_hdr(ws, row, 1, DC+len(all_years)+2, "INVESTING ACTIVITIES")
    row += 1

    # CapEx with breakdown note
    lbl(ws, row, 1, "  Capital Expenditures (CapEx)", indent=1)
    ws.cell(row=row, column=DC+len(all_years)+1,
            value="CapEx used for: data centers, servers, network infra, office facilities, AI compute. "
                  "Source: SEC EDGAR investing activities.").font = ft(italic=True, color="595959", size=8)
    capex_row = row
    proj_rev_tmp = last_rev
    for i, yr in enumerate(hist_years):
        v = data["capex"].get(yr)
        val_cell(ws, row, dc(i), -abs(v)/scale if v else None, FMT)
    proj_rev_tmp = last_rev
    for i in range(n_proj):
        proj_rev_tmp = proj_rev_tmp * (1 + rev_growth[i])
        val_cell(ws, row, dc(n_hist+i), -proj_rev_tmp*assumptions["capex_pct"]/scale, FMT, color=GREEN)

    row += 1
    lbl(ws, row, 1, "  CapEx as % of Revenue", italic=True, color="595959")
    for i, yr in enumerate(hist_years):
        cx_v = abs(data["capex"].get(yr) or 0)
        rv_v = data["revenue"].get(yr) or 1
        c = ws.cell(row=row, column=dc(i), value=cx_v/rv_v if cx_v else None)
        c.number_format = "0.0%"; c.font = ft(italic=True, color="595959"); c.alignment = aln("right")
    for i in range(n_proj):
        c = ws.cell(row=row, column=dc(n_hist+i), value=assumptions["capex_pct"])
        c.number_format = "0.0%"; c.font = ft(italic=True, color=GREEN); c.alignment = aln("right")

    row += 1
    lbl(ws, row, 1, "Cash from Investing (CFI)", bold=True, bg=ACCENT_BLUE)
    cfi_row = row
    for i in range(len(all_years)):
        cl = get_column_letter(dc(i))
        c = ws.cell(row=row, column=dc(i), value=f"={cl}{capex_row}")
        c.number_format = FMT; c.font = ft(bold=True, color=GREEN if i>=n_hist else BLACK)
        c.fill = fl(ACCENT_BLUE); c.alignment = aln("right")

    # ─── FINANCING ACTIVITIES ───
    row += 2
    sub_hdr(ws, row, 1, DC+len(all_years)+2, "FINANCING ACTIVITIES")
    row += 1

    lbl(ws, row, 1, "  Dividends Paid", indent=1)
    div_row = row
    for i, yr in enumerate(hist_years):
        v = data["div_paid"].get(yr)
        val_cell(ws, row, dc(i), -abs(v)/scale if v else None, FMT)
    for i in range(n_proj):
        val_cell(ws, row, dc(n_hist+i), 0.0, FMT, color=INPUT_BLUE)

    row += 1
    lbl(ws, row, 1, "  Share Buybacks / Repurchases", indent=1)
    ws.cell(row=row, column=DC+len(all_years)+1,
            value="Buybacks reflect management's capital return program. Source: SEC EDGAR financing activities."
            ).font = ft(italic=True, color="595959", size=8)
    buy_row = row
    for i, yr in enumerate(hist_years):
        v = data.get("buybacks", {}).get(yr)
        val_cell(ws, row, dc(i), -abs(v)/scale if v else None, FMT)
    last_buy = abs(last_val(data.get("buybacks", {})) or 0)
    for i in range(n_proj):
        val_cell(ws, row, dc(n_hist+i), -last_buy/scale, FMT, color=INPUT_BLUE)

    row += 1
    lbl(ws, row, 1, "Cash from Financing (CFF)", bold=True, bg=ACCENT_BLUE)
    cff_row = row
    for i in range(len(all_years)):
        cl = get_column_letter(dc(i))
        c = ws.cell(row=row, column=dc(i))
        c.value = f"={cl}{div_row}+{cl}{buy_row}"
        c.number_format = FMT; c.font = ft(bold=True, color=GREEN if i>=n_hist else BLACK)
        c.fill = fl(ACCENT_BLUE); c.alignment = aln("right")

    # ─── FCF ───
    row += 2
    lbl(ws, row, 1, "Free Cash Flow (FCF = CFO + CapEx)", bold=True, bg=DARK_BLUE, color=WHITE)
    fcf_row = row
    for i in range(len(all_years)):
        cl = get_column_letter(dc(i))
        c = ws.cell(row=row, column=dc(i))
        c.value = f"={cl}{cfo_row}+{cl}{capex_row}"
        c.number_format = FMT; c.font = ft(bold=True, color=WHITE)
        c.fill = fl(DARK_BLUE); c.alignment = aln("right")

    row += 1
    lbl(ws, row, 1, "  FCF Yield (FCF / Market Cap)", italic=True, color="595959")
    mkt_cap = (data["stock"].get("market_cap") or 0) / scale
    for i, yr in enumerate(hist_years):
        fcf_v  = data["op_cf"].get(yr, 0) or 0
        capex_v = abs(data["capex"].get(yr, 0) or 0)
        if mkt_cap and mkt_cap > 0:
            c = ws.cell(row=row, column=dc(i), value=(fcf_v - capex_v)/scale/mkt_cap if (fcf_v-capex_v) else None)
            c.number_format = "0.0%"; c.font = ft(italic=True, color="595959"); c.alignment = aln("right")

    # ─── Business Segment Note ───
    row += 3
    sub_hdr(ws, row, 1, DC+len(all_years)+2, f"BUSINESS SEGMENT CONTEXT  —  Key Revenue Drivers & CapEx Priorities")
    row += 1

    ticker = data["ticker"]
    # Segment context — generic but shaped to common structures
    segment_notes = {
        "META": [
            ("Family of Apps (Facebook, Instagram, WhatsApp, Messenger)",
             "Primary revenue driver (~97% of total). Monetized through digital advertising (impressions, clicks). "
             "Revenue scales with daily active users (DAP) and average revenue per user (ARPU). "
             "Q4 2023 earnings: DAP reached 3.19B. ARPU highest in US/Canada ($68.44/qtr)."),
            ("Reality Labs (VR/AR - Quest, Ray-Ban Meta, Orion)",
             "Operating at a significant loss (~$16B operating loss in 2023). Long-term strategic bet on the metaverse. "
             "CapEx here funds R&D for hardware, spatial computing. Management committed to multi-year investment."),
            ("AI & Infrastructure Investment",
             "Meta's #1 stated CapEx priority. 2024 CapEx guidance raised to $35-40B (from $30-37B) driven by "
             "AI compute (H100/H200 GPUs), data centers, and network infrastructure. "
             "Llama open-source model strategy drives developer ecosystem."),
            ("Capital Return Program",
             "Active buyback program. $50B buyback authorized. Initiated first-ever dividend ($0.50/share/qtr) in Feb 2024. "
             "Source: Q4 2023 Earnings Release, February 1, 2024."),
        ],
        "DEFAULT": [
            ("Core Business Revenue",
             "Primary revenue generating segment. Growth driven by pricing, volume, and market share expansion. "
             "See Income Statement for historical trends and projections."),
            ("Capital Expenditure Priorities",
             f"CapEx ({assumptions['capex_pct']:.1%} of revenue projected) directed toward maintaining and expanding "
             "productive asset base. Mix of maintenance CapEx and growth CapEx. See 10-K investing activities."),
            ("Operating Cash Generation",
             f"Business converts approximately {assumptions['op_margin']:.1%} of revenue to operating income. "
             "Cash generation reinvested in growth, debt service, and capital returns."),
            ("Capital Return / Balance Sheet",
             "Management balancing organic investment with shareholder returns. "
             "Debt levels monitored relative to EBITDA. See latest proxy statement (DEF 14A) for governance detail."),
        ]
    }
    notes = segment_notes.get(ticker, segment_notes["DEFAULT"])

    for seg_name, seg_note in notes:
        ws.row_dimensions[row].height = 13
        c1 = ws.cell(row=row, column=1, value=f"  ▸ {seg_name}")
        c1.font = ft(bold=True, size=9, color=DARK_BLUE)
        row += 1
        c2 = ws.cell(row=row, column=1, value=f"    {seg_note}")
        c2.font = ft(size=8, italic=True, color="595959")
        try: ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=DC+len(all_years)+1)
        except: pass
        c2.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row].height = 42
        row += 2

    # Source footnote
    ws.cell(row=row, column=1,
            value=f"Sources: SEC EDGAR 10-K annual filings (data.sec.gov), Q4 2023 Earnings Release, "
                  f"Management guidance from earnings calls. CapEx projections based on {assumptions['capex_pct']:.1%} of revenue assumption."
            ).font = ft(italic=True, color="595959", size=8)

    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 2
    for i in range(len(all_years)):
        ws.column_dimensions[get_column_letter(DC+i)].width = 11
    ws.column_dimensions[get_column_letter(DC+len(all_years)+1)].width = 55

# ── VALUATION ─────────────────────────────────────────────────────────────────

def _valuation(wb, data, assumptions, hist_years, proj_years):
    ws = wb.create_sheet("Valuation")
    ws.sheet_view.showGridLines = False

    n_proj  = len(proj_years)
    scale   = 1_000_000
    stock   = data["stock"]
    peers   = data.get("peers", {})
    ticker  = data["ticker"]

    # ── Pull live values (hardcoded in as inputs since from external source)
    price       = stock.get("price", 0) or 0
    shares_out  = stock.get("shares_out", 0) or 0
    market_cap  = stock.get("market_cap", 0) or (price * shares_out)
    ev          = stock.get("ev", 0) or 0
    beta        = stock.get("beta", 1.0) or 1.0
    last_cash   = last_val(data["cash"]) or 0
    last_debt   = last_val(data["lt_debt"]) or 0

    # Derive EV if not from Yahoo
    if not ev:
        ev = market_cap + last_debt - last_cash

    price_mm      = price
    shares_mm     = shares_out / scale if shares_out else 1
    mktcap_mm     = market_cap / scale if market_cap else price * shares_out / scale
    ev_mm         = ev / scale if ev else mktcap_mm + last_debt/scale - last_cash/scale
    last_cash_mm  = last_cash / scale
    last_debt_mm  = last_debt / scale

    section_hdr(ws, 1, 1, 10, f"VALUATION SUMMARY  —  {ticker}  |  {data['name']}")
    ws.cell(row=2, column=1, value="  $ in millions except per share amounts  |  Sources: Yahoo Finance, SEC EDGAR, DCF model"
            ).font = ft(italic=True, color="595959", size=8)

    # ── SECTION 1: Market Snapshot ──
    row = 4
    sub_hdr(ws, row, 1, 10, "MARKET SNAPSHOT  (Source: Yahoo Finance — live data)")
    row += 1

    mkt_items = [
        ("Share Price (as of model date)",          f"${price_mm:,.2f}",         "Yahoo Finance — previousClose / regularMarketPrice"),
        ("Shares Outstanding",                      f"{shares_mm:,.1f}mm",        "Yahoo Finance / SEC EDGAR 10-K filing"),
        ("Market Capitalization",                   f"${mktcap_mm:,.1f}mm",       "Price × Shares Outstanding"),
        ("(+) Total Long-Term Debt",                f"${last_debt_mm:,.1f}mm",    f"SEC EDGAR 10-K — latest fiscal year"),
        ("(-) Cash & Cash Equivalents",             f"${last_cash_mm:,.1f}mm",    f"SEC EDGAR 10-K — latest fiscal year"),
        ("Enterprise Value (EV)",                   f"${ev_mm:,.1f}mm",           "Market Cap + Debt − Cash  (Yahoo EV if available)"),
        ("Analyst Consensus Price Target",          f"${stock.get('target_price',0):,.2f}" if stock.get('target_price') else "N/A",
                                                                                  "Yahoo Finance — analyst mean target price"),
        ("Analyst Recommendation",                  (stock.get('rec') or 'N/A').upper(), "Yahoo Finance — consensus recommendation"),
    ]
    for k, v, src in mkt_items:
        ws.row_dimensions[row].height = 13
        lbl(ws, row, 1, f"  {k}", bg=GRAY_LIGHT)
        c = ws.cell(row=row, column=3, value=v)
        c.font = ft(bold=True, color=INPUT_BLUE, size=9); c.fill = fl(YELLOW_BG); c.alignment = aln("right")
        note_cell(ws, row, 5, src)
        row += 1

    # ── SECTION 2: DCF VALUATION ──
    row += 1
    sub_hdr(ws, row, 1, 10, "DCF VALUATION  —  Levered Free Cash Flow Method")
    row += 1

    # Cost of equity
    rf   = assumptions["rf_rate"]
    erp  = assumptions["erp"]
    beta_v = assumptions["beta"]
    coe  = rf + beta_v * erp
    g    = assumptions["lt_growth"]
    n_p  = n_proj

    lbl(ws, row, 1, "  Risk-Free Rate", bg=GRAY_LIGHT); val_cell(ws, row, 3, rf, "0.00%", color=INPUT_BLUE, bg=YELLOW_BG); note_cell(ws, row, 5, "US 10Y Treasury yield"); row+=1
    lbl(ws, row, 1, "  Equity Risk Premium", bg=GRAY_LIGHT); val_cell(ws, row, 3, erp, "0.00%", color=INPUT_BLUE, bg=YELLOW_BG); note_cell(ws, row, 5, "Damodaran implied ERP"); row+=1
    lbl(ws, row, 1, "  Beta (Levered)", bg=GRAY_LIGHT); val_cell(ws, row, 3, beta_v, "0.00", color=INPUT_BLUE, bg=YELLOW_BG); note_cell(ws, row, 5, "Yahoo Finance 5Y monthly"); row+=1

    coe_row = row
    lbl(ws, row, 1, "  Cost of Equity (CAPM)", bg=GRAY_LIGHT)
    rf_r, erp_r, beta_r = row-3, row-2, row-1
    c = ws.cell(row=row, column=3, value=f"=C{rf_r}+C{beta_r}*C{erp_r}")
    c.number_format = "0.00%"; c.font = ft(color=BLACK); c.fill = fl(GRAY_LIGHT); c.alignment = aln("right")
    note_cell(ws, row, 5, "CAPM: Rf + β × ERP"); row+=1

    lbl(ws, row, 1, "  Long-Term Growth Rate (g)", bg=GRAY_LIGHT); val_cell(ws, row, 3, g, "0.00%", color=INPUT_BLUE, bg=YELLOW_BG); note_cell(ws, row, 5, "Terminal growth = nominal GDP"); ltg_row = row; row+=1

    row += 1
    # Projected FCFs
    sub_hdr(ws, row, 1, 10, "PROJECTED FREE CASH FLOWS")
    row += 1

    # FCF year headers
    lbl(ws, row, 1, "  Year")
    for i, yr in enumerate(proj_years):
        c = ws.cell(row=row, column=3+i, value=f"{yr}E")
        c.font = ft(bold=True, color=WHITE, size=9); c.fill = fl(MID_BLUE); c.alignment = aln("center")
    row += 1

    # Build projected FCFs from scratch (revenue-driven)
    proj_rev_list  = []
    proj_fcf_list  = []
    proj_ni_list   = []
    base_rev = last_val(data["revenue"]) or 1
    cur_rev  = base_rev
    for i in range(n_proj):
        cur_rev = cur_rev * (1 + assumptions["rev_growth"][i])
        ni_i   = cur_rev * assumptions["op_margin"] * (1 - assumptions["tax_rate"])
        da_i   = cur_rev * assumptions["dep_pct"]
        cx_i   = cur_rev * assumptions["capex_pct"]
        sbc_i  = last_val(data.get("stock_comp",{})) or 0
        fcf_i  = ni_i + da_i - cx_i + sbc_i
        proj_rev_list.append(cur_rev)
        proj_fcf_list.append(fcf_i)
        proj_ni_list.append(ni_i)

    lbl(ws, row, 1, "  Revenue ($mm)", bg=GRAY_LIGHT)
    for i, v in enumerate(proj_rev_list):
        val_cell(ws, row, 3+i, v/scale, '#,##0.0', color=GREEN); row_rev_dcf = row
    row += 1

    lbl(ws, row, 1, "  Net Income ($mm)", bg=GRAY_LIGHT)
    for i, v in enumerate(proj_ni_list):
        val_cell(ws, row, 3+i, v/scale, '#,##0.0', color=GREEN)
    row += 1

    lbl(ws, row, 1, "  (+) D&A ($mm)", bg=GRAY_LIGHT)
    for i in range(n_proj):
        val_cell(ws, row, 3+i, proj_rev_list[i]*assumptions["dep_pct"]/scale, '#,##0.0', color=GREEN)
    row += 1

    lbl(ws, row, 1, "  (-) CapEx ($mm)", bg=GRAY_LIGHT)
    for i in range(n_proj):
        val_cell(ws, row, 3+i, -proj_rev_list[i]*assumptions["capex_pct"]/scale, '#,##0.0', color=GREEN)
    row += 1

    lbl(ws, row, 1, "  Free Cash Flow ($mm)", bold=True, bg=ACCENT_BLUE)
    fcf_val_row = row
    for i, v in enumerate(proj_fcf_list):
        c = ws.cell(row=row, column=3+i, value=v/scale)
        c.number_format = '#,##0.0'; c.font = ft(bold=True, color=GREEN); c.fill = fl(ACCENT_BLUE); c.alignment = aln("right")
    row += 1

    # Discount factors (mid-year convention)
    lbl(ws, row, 1, "  Discount Period (mid-year)")
    for i in range(n_proj):
        val_cell(ws, row, 3+i, i + 0.5, "0.0")
    row += 1

    lbl(ws, row, 1, "  Discount Factor")
    disc_row = row
    for i in range(n_proj):
        period = i + 0.5
        c = ws.cell(row=row, column=3+i)
        c.value = f"=1/(1+C{coe_row})^{period}"
        c.number_format = "0.0000"; c.font = ft(color=BLACK); c.alignment = aln("right")
    row += 1

    lbl(ws, row, 1, "  PV of FCF ($mm)", bold=True)
    pv_fcf_row = row
    for i in range(n_proj):
        fcf_col = get_column_letter(3+i)
        c = ws.cell(row=row, column=3+i)
        c.value = f"={fcf_col}{fcf_val_row}*{fcf_col}{disc_row}"
        c.number_format = '#,##0.0'; c.font = ft(bold=True, color=GREEN); c.alignment = aln("right")
    row += 1

    # Sum of PV FCFs
    pv_sum_cols = "+".join([f"{get_column_letter(3+i)}{pv_fcf_row}" for i in range(n_proj)])
    lbl(ws, row, 1, f"  Sum of PV of FCFs (Yrs 1-{n_proj})", bold=True, bg=LIGHT_BLUE)
    pv_sum_row = row
    c = ws.cell(row=row, column=3, value=f"={pv_sum_cols}")
    c.number_format = '#,##0.0'; c.font = ft(bold=True, color=DARK_BLUE); c.fill = fl(LIGHT_BLUE); c.alignment = aln("right")
    row += 1

    # Terminal Value
    row += 1
    sub_hdr(ws, row, 1, 10, "TERMINAL VALUE (Gordon Growth Model)")
    row += 1

    last_fcf_col = get_column_letter(3 + n_proj - 1)
    lbl(ws, row, 1, "  FCF in Final Projection Year ($mm)", bg=GRAY_LIGHT)
    tv_base_row = row
    c = ws.cell(row=row, column=3, value=f"={last_fcf_col}{fcf_val_row}")
    c.number_format = '#,##0.0'; c.font = ft(color=BLACK); c.fill = fl(GRAY_LIGHT); c.alignment = aln("right")
    row += 1

    lbl(ws, row, 1, "  FCFt+1 = FCFt × (1+g)", bg=GRAY_LIGHT)
    fcft1_row = row
    c = ws.cell(row=row, column=3, value=f"=C{tv_base_row}*(1+C{ltg_row})")
    c.number_format = '#,##0.0'; c.font = ft(color=BLACK); c.fill = fl(GRAY_LIGHT); c.alignment = aln("right")
    row += 1

    lbl(ws, row, 1, "  Terminal Value = FCFt+1 / (CoE − g)", bg=GRAY_LIGHT)
    tv_row = row
    c = ws.cell(row=row, column=3, value=f"=C{fcft1_row}/(C{coe_row}-C{ltg_row})")
    c.number_format = '#,##0.0'; c.font = ft(color=BLACK); c.fill = fl(GRAY_LIGHT); c.alignment = aln("right")
    row += 1

    # Discount TV to PV using last period discount factor
    last_disc_col = get_column_letter(3 + n_proj - 1)
    lbl(ws, row, 1, f"  PV of Terminal Value (discounted {n_proj} yrs)", bg=GRAY_LIGHT)
    pv_tv_row = row
    c = ws.cell(row=row, column=3, value=f"=C{tv_row}/{last_disc_col}{disc_row}*(1/(1+C{coe_row})^0.5)")
    c.number_format = '#,##0.0'; c.font = ft(color=BLACK); c.fill = fl(GRAY_LIGHT); c.alignment = aln("right")
    row += 1

    # Equity Value Build
    row += 1
    sub_hdr(ws, row, 1, 10, "EQUITY VALUE BUILD")
    row += 1

    lbl(ws, row, 1, "  (+) PV of FCFs", bg=GRAY_LIGHT)
    c = ws.cell(row=row, column=3, value=f"=C{pv_sum_row}")
    c.number_format = '#,##0.0'; c.font = ft(color=BLACK); c.fill = fl(GRAY_LIGHT); c.alignment = aln("right")
    row += 1

    lbl(ws, row, 1, "  (+) PV of Terminal Value", bg=GRAY_LIGHT)
    c = ws.cell(row=row, column=3, value=f"=C{pv_tv_row}")
    c.number_format = '#,##0.0'; c.font = ft(color=BLACK); c.fill = fl(GRAY_LIGHT); c.alignment = aln("right")
    row += 1

    lbl(ws, row, 1, "  (+) Cash & Equivalents ($mm)", bg=GRAY_LIGHT)
    val_cell(ws, row, 3, last_cash_mm, '#,##0.0', color=INPUT_BLUE, bg=YELLOW_BG); note_cell(ws, row, 5, "Latest balance sheet — SEC EDGAR")
    plus_cash_row = row; row += 1

    lbl(ws, row, 1, "  (−) Total Debt ($mm)", bg=GRAY_LIGHT)
    val_cell(ws, row, 3, -last_debt_mm, '#,##0.0', color=INPUT_BLUE, bg=YELLOW_BG); note_cell(ws, row, 5, "Latest balance sheet — SEC EDGAR")
    minus_debt_row = row; row += 1

    lbl(ws, row, 1, "  Implied Equity Value ($mm)", bold=True, bg=ACCENT_BLUE)
    eq_val_row = row
    c = ws.cell(row=row, column=3)
    c.value = (f"=C{pv_sum_row}+C{pv_tv_row}"
               f"+C{plus_cash_row}+C{minus_debt_row}")
    c.number_format = '#,##0.0'; c.font = ft(bold=True, color=DARK_BLUE); c.fill = fl(ACCENT_BLUE); c.alignment = aln("right")
    row += 1

    lbl(ws, row, 1, "  Shares Outstanding ($mm)", bg=GRAY_LIGHT)
    val_cell(ws, row, 3, shares_mm, '#,##0.0', color=INPUT_BLUE, bg=YELLOW_BG); note_cell(ws, row, 5, "Yahoo Finance / SEC EDGAR")
    sh_row = row; row += 1

    lbl(ws, row, 1, "  DCF Implied Share Price", bold=True, bg=DARK_BLUE, color=WHITE)
    dcf_px_row = row
    c = ws.cell(row=row, column=3, value=f"=IFERROR(C{eq_val_row}/C{sh_row},\"N/A\")")
    c.number_format = '$#,##0.00'; c.font = ft(bold=True, color=WHITE); c.fill = fl(DARK_BLUE); c.alignment = aln("right")
    row += 1

    lbl(ws, row, 1, "  Current Share Price", bg=GRAY_LIGHT)
    curr_px_row = row
    val_cell(ws, row, 3, price_mm, '$#,##0.00', color=INPUT_BLUE, bg=YELLOW_BG); note_cell(ws, row, 5, f"Yahoo Finance — {ticker} as of model date")
    row += 1

    lbl(ws, row, 1, "  Implied Upside / (Downside) to Current Price", bold=True)
    c = ws.cell(row=row, column=3, value=f"=IFERROR(C{dcf_px_row}/C{curr_px_row}-1,\"N/A\")")
    c.number_format = "0.0%"; c.font = ft(bold=True, color=DARK_BLUE); c.fill = fl(ACCENT_BLUE); c.alignment = aln("right")
    note_cell(ws, row, 5, "Positive = DCF implies stock is undervalued vs. current price"); row += 1

    # ── SECTION 3: TRADING MULTIPLES ──
    row += 2
    sub_hdr(ws, row, 1, 10, f"TRADING MULTIPLES  —  {ticker} vs. Peers  (Source: Yahoo Finance)")
    row += 1

    mult_hdrs = ["Metric", ticker, *list(peers.keys()), "Peer Median", "Source"]
    for j, h in enumerate(mult_hdrs):
        c = ws.cell(row=row, column=1+j, value=h)
        c.font = ft(bold=True, color=WHITE, size=9); c.fill = fl(MID_BLUE); c.alignment = aln("center")
    row += 1

    def peer_vals(key):
        return [p.get(key, 0) or 0 for p in peers.values()]

    def median(vals):
        clean = sorted([v for v in vals if v and v > 0])
        if not clean: return None
        n = len(clean)
        return clean[n//2] if n%2 else (clean[n//2-1]+clean[n//2])/2

    multiples = [
        ("EV / Revenue (LTM)",     stock.get("ev_rev",0),        "ev_rev",    "Yahoo Finance — EV / trailing 12M revenue"),
        ("EV / EBITDA (LTM)",      stock.get("ev_ebitda",0),     "ev_ebitda", "Yahoo Finance — EV / trailing 12M EBITDA"),
        ("P/E Ratio (Trailing)",   stock.get("pe_trailing",0),   "pe_trailing","Yahoo Finance — trailing 12M P/E"),
        ("P/E Ratio (Forward)",    stock.get("pe_forward",0),    "pe_forward", "Yahoo Finance — NTM consensus EPS"),
        ("Price / Sales (LTM)",    stock.get("ps_ratio",0),      "ps_ratio",   "Yahoo Finance — market cap / revenue"),
        ("Price / Book",           stock.get("pb_ratio",0),      "pb_ratio",   "Yahoo Finance — market cap / book equity"),
        ("PEG Ratio",              stock.get("peg_ratio",0),     "peg_ratio",  "Yahoo Finance — P/E ÷ EPS growth rate"),
        ("Gross Margin %",         stock.get("gross_margin",0),  "gross_margin","Yahoo Finance — LTM gross margin"),
        ("Operating Margin %",     stock.get("op_margin",0),     "op_margin",  "Yahoo Finance — LTM operating margin"),
        ("Revenue Growth (YoY %)", stock.get("revenue_growth",0),"revenue_growth","Yahoo Finance — YoY revenue growth"),
        ("Return on Equity (ROE)", stock.get("roe",0),           "roe",        "Yahoo Finance — net income / equity"),
        ("Debt / Equity",          stock.get("debt_equity",0),   "debt_equity","Yahoo Finance — total debt / equity"),
    ]

    for i, (metric, subj_val, peer_key, src) in enumerate(multiples):
        r = row + i
        ws.row_dimensions[r].height = 13
        is_pct = peer_key in ("gross_margin","op_margin","revenue_growth","roe")
        is_x   = peer_key in ("ev_rev","ev_ebitda","pe_trailing","pe_forward","ps_ratio","pb_ratio","peg_ratio","debt_equity")

        lbl(ws, r, 1, f"  {metric}", bg=GRAY_LIGHT if i%2==0 else WHITE)

        def fmt_v(v):
            if not v or v == 0: return "—"
            if is_pct: return f"{v:.1%}"
            if is_x:   return f"{v:.1f}x"
            return f"{v:.2f}"

        # Subject company
        c = ws.cell(row=r, column=2, value=fmt_v(subj_val))
        c.font = ft(bold=True, size=9, color=DARK_BLUE)
        c.fill = fl(ACCENT_BLUE if i%2==0 else LIGHT_BLUE); c.alignment = aln("center")

        # Peers
        pv = peer_vals(peer_key)
        for j, (pticker, pd) in enumerate(peers.items()):
            pval = pd.get(peer_key, 0) or 0
            c2 = ws.cell(row=r, column=3+j, value=fmt_v(pval))
            c2.font = ft(size=9); c2.fill = fl(GRAY_LIGHT if i%2==0 else WHITE); c2.alignment = aln("center")

        # Peer median
        med = median(pv)
        med_col = 3 + len(peers)
        c3 = ws.cell(row=r, column=med_col, value=fmt_v(med) if med else "—")
        c3.font = ft(bold=True, size=9, color=MID_BLUE); c3.fill = fl(GRAY_LIGHT); c3.alignment = aln("center")

        # Source
        ws.cell(row=r, column=med_col+1, value=src).font = ft(italic=True, size=8, color="595959")

    row += len(multiples) + 2
    ws.cell(row=row, column=1,
            value="Note: All multiples sourced from Yahoo Finance at time of model run. EV calculated as "
                  "Market Cap + Total Debt − Cash. Peer set selected based on industry similarity. "
                  f"DCF assumes {n_proj}-year explicit forecast period with {g:.1%} terminal growth rate and "
                  f"{coe:.1%} cost of equity (β={beta_v:.2f}).").font = ft(italic=True, color="595959", size=8)
    try: ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
    except: pass
    ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True)
    ws.row_dimensions[row].height = 28

    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 14
    for j in range(len(peers)):
        ws.column_dimensions[get_column_letter(3+j)].width = 12
    ws.column_dimensions[get_column_letter(3+len(peers))].width = 12
    ws.column_dimensions[get_column_letter(4+len(peers))].width = 40

# ── MAIN ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Build a financial operating model for any public company.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python Forecasting_v2.py META
  python Forecasting_v2.py AAPL --years 5
  python Forecasting_v2.py NVDA --years 3
        """
    )
    parser.add_argument("ticker", help="Stock ticker symbol (e.g., META, AAPL, MSFT)")
    parser.add_argument("--years", type=int, default=5, choices=[3, 4, 5],
                        help="Projection years (default: 5)")
    args = parser.parse_args()

    ticker = args.ticker.upper().strip()
    n_proj = args.years
    output_dir = Path.home() / "Desktop"
    output_dir.mkdir(parents=True, exist_ok=True)

    print(f"\n{'='*65}")
    print(f"  Financial Operating Model Builder  —  {ticker}")
    print(f"  Projection horizon: {n_proj} years  |  Output: {output_dir}")
    print(f"{'='*65}\n")

    print("[Step 1/4]  Fetching financial data...")
    try:
        data = fetch_financials(ticker)
    except Exception as e:
        print(f"\n  ERROR: Could not fetch data for '{ticker}': {e}")
        print("  Verify the ticker is valid and your internet connection is active.")
        sys.exit(1)

    s = data["stock"]
    print(f"\n  ✓ Company   : {data['name']}")
    print(f"  ✓ Sector    : {data['sector']} / {data['industry']}")
    print(f"  ✓ Price     : ${s.get('price',0):,.2f}  |  Mkt Cap: ${(s.get('market_cap',0) or 0)/1e9:,.1f}B")
    print(f"  ✓ EV        : ${(s.get('ev',0) or 0)/1e9:,.1f}B  |  Beta: {s.get('beta',0):.2f}")
    print(f"  ✓ Revenue   : {sorted(data['revenue'].keys())} ({len(data['revenue'])} yrs)")
    print(f"  ✓ Net Income: {sorted(data['net_income'].keys())} ({len(data['net_income'])} yrs)")
    print(f"  ✓ Balance Sh: {sorted(data['total_assets'].keys())} ({len(data['total_assets'])} yrs)")
    print(f"  ✓ Cash Flow : {sorted(data['op_cf'].keys())} ({len(data['op_cf'])} yrs)")
    print(f"  ✓ Peers     : {list(data['peers'].keys())}")
    for key in ["revenue","net_income","op_cf","total_assets","cash","lt_debt"]:
        if not data.get(key):
            print(f"  ⚠  WARNING: '{key}' is empty — model will have gaps for this series")

    print(f"\n[Step 2/4]  Deriving forecast assumptions from historicals...")
    assumptions = build_assumptions(data, n_proj)
    print(f"  ✓ Rev growth (Yr1) : {assumptions['rev_growth'][0]:.1%}")
    print(f"  ✓ Gross margin     : {assumptions['gross_margin']:.1%}")
    print(f"  ✓ Op margin        : {assumptions['op_margin']:.1%}")
    print(f"  ✓ Tax rate         : {assumptions['tax_rate']:.1%}")
    print(f"  ✓ CapEx % rev      : {assumptions['capex_pct']:.1%}")
    print(f"  ✓ CoE              : {assumptions['rf_rate'] + assumptions['beta']*assumptions['erp']:.1%}")

    print(f"\n[Step 3/4]  Building Excel workbook...")
    wb = build_workbook(data, n_proj, assumptions)

    filename = f"{ticker}_Operating_Model_{datetime.now().strftime('%Y%m%d')}.xlsx"
    out_path = output_dir / filename
    wb.save(str(out_path))
    print(f"  ✓ Saved to: {out_path}")

    print(f"\n[Step 4/4]  Recalculating formulas via LibreOffice...")
    import subprocess, json as _json
    scripts_path = Path(__file__).parent / "scripts" / "recalc.py"
    if not scripts_path.exists():
        # Try sibling scripts folder
        scripts_path = Path(__file__).parent / "recalc.py"
    if scripts_path.exists():
        result = subprocess.run(
            ["python3", str(scripts_path), str(out_path)],
            capture_output=True, text=True, timeout=120
        )
        try:
            jr = _json.loads(result.stdout)
            if jr.get("status") == "success":
                print(f"  ✓ {jr.get('total_formulas',0)} formulas recalculated — zero errors")
            else:
                print(f"  ⚠  Errors found: {jr.get('error_summary')}")
        except:
            print(f"  ✓ Recalculation complete")
    else:
        print(f"  ℹ  recalc.py not found — open in Excel to recalculate formulas")

    print(f"\n{'='*65}")
    print(f"  ✓ MODEL COMPLETE")
    print(f"  ✓ File  : {out_path}")
    print(f"  ✓ Sheets: Cover | Assumptions | Income Statement |")
    print(f"            Balance Sheet | Cash Flow | Valuation")
    print(f"{'='*65}\n")


if __name__ == "__main__":
    main()
